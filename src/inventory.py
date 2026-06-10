"""src/inventory.py — Inventory analytics (ERP-Stock-Balance aligned).

Stock source: MsItemBatchOpening (live snapshot) + TrVocItem movements
for historical roll-back, with the same FY-CASE filter we use everywhere
to drop duplicate FY-tagged rows.

Cost source: MsItemMaster.ValuationCaseRate — the ERP's maintained
landed-cost rate that already includes state excise for Daman variants
(verified to match ERP stock-value report within ~1%).

Movement classification: MsTransType.QtyInOut (I = inward, O = outward).
"""
from __future__ import annotations

import calendar
import json
import os
import re
import sys
from datetime import date, timedelta

import pandas as pd
import plotly.graph_objects as go
import streamlit as st

from db import run_query
from utils.helpers import (format_inr, CASES_SQL_EXPR as _CASES, safe_section,
                          cases_sql, keg_mode_toggle)

# ERP Stock & Sale baseline — anchors live stock to the ERP's authoritative
# closing, then rolls forward with live movements (the ERP's per-item opening
# stock isn't stored in any accessible table, so we can't compute it purely
# live; this baseline is refreshed by dropping in a newer S&S export).
_BASELINE_PATH = os.path.join(os.path.dirname(__file__), "..", "data",
                              "stock_baseline.json")


def _safe_format(df: "pd.DataFrame", fmt: dict) -> dict:
    """Drop any keys from the format dict that aren't columns of `df`.

    Defensive guard: prevents KeyError crashes when a column is renamed
    or absent. Pair with df.style.format(_safe_format(df, fmt)).
    """
    return {k: v for k, v in fmt.items() if k in df.columns}

PURCHASE_TYPES: tuple[int, ...] = (11, 20, 22, 30, 32, 33, 36, 42, 45, 46, 48, 54)
IMPORT_TYPES:   tuple[int, ...] = (22, 54)               # imports proper
DAMAN_TYPES:    tuple[int, ...] = (42,)                  # Daman / cross-state
SALES_TYPES:    tuple[int, ...] = (18, 19, 23, 35, 37, 38, 39, 40, 41, 44, 47, 49, 51, 53)

# ── Brindco transition (Jun-2026) ───────────────────────────────────────────
# The 6 Diageo brands moving to Brindco. Each entry: (code, display name,
# BrandName LIKE patterns, annual target in cases). Patterns roll up every
# variant (NEW / Hipster / 6-pack re-codes) under the same consumer brand.
BRINDCO_PRESET: list[dict] = [
    {"code": "CAOL",  "name": "Caol Ila 12 YO",
     "patterns": ["CAOL"],                        "target": 75},
    {"code": "CIRV",  "name": "Cîroc Vodka",
     "patterns": ["CIROC"],                       "target": 200},
    {"code": "DALW",  "name": "Dalwhinnie 15 YO",
     "patterns": ["DALWHINNIE", "DAL WHINNIE"],   "target": 75},
    {"code": "GORD",  "name": "Gordon's London Dry",
     "patterns": ["GORDON"],                      "target": 2000},
    {"code": "J&B",   "name": "J & B Rare",
     "patterns": ["J & B", "J&B"],                "target": 4000},
    {"code": "TA10S", "name": "Talisker 10 YO",
     "patterns": ["TALISKER"],                    "target": 500},
]
# "Below this we treat the month as a stock-out, not real demand" — per the
# owner's rule (a 0.17-case month is a dribble, not market signal).
_STOCKOUT_THRESHOLD_CS: float = 1.0


def _month_first(d: date, months_offset: int = 0) -> date:
    m = d.month - 1 + months_offset
    return date(d.year + m // 12, m % 12 + 1, 1)

_PRINCIPAL_NAMES: dict[str, str] = {
    "C00025": "United Spirits",
    "C00040": "Diageo",
    "C00039": "United Breweries",
    "C00056": "Brown-Forman",
}
_PRINCIPAL_COLOR: dict[str, str] = {
    "United Spirits":   "#1B4F72",
    "Diageo":           "#378ADD",
    "United Breweries": "#1D9E75",
    "Brown-Forman":     "#EF9F27",
}
_KPI_COLORS = ["#1B4F72", "#378ADD", "#1D9E75", "#EF9F27"]

# FY-CASE join fragment that kills TrVocItem duplicate FY rows.
_FY_JOIN = """
    AND vi.FinancialYear = CASE
        WHEN MONTH(h.VoucherDate) >= 4
        THEN CAST(YEAR(h.VoucherDate) AS VARCHAR)
             + '-' + CAST(YEAR(h.VoucherDate)+1 AS VARCHAR)
        ELSE CAST(YEAR(h.VoucherDate)-1 AS VARCHAR)
             + '-' + CAST(YEAR(h.VoucherDate) AS VARCHAR)
    END
"""


# ═══════════════════════════════════════════════════════════════════════════════
# DATA LOADERS
# ═══════════════════════════════════════════════════════════════════════════════

def _kegaware_cases(desc: str, units: float, bpc) -> float:
    """Convert a physical quantity (bottles, or keg count) to case-equivalents
    using the SAME keg rule as CASES_SQL_EXPR (20LT=2.56, 30LT=3.85, 50LT=6.41
    cases; everything else = bottles / BottlesPerCase). Keeps the Inventory
    page's case counts consistent with Purchase / Sales / Sales-Plan."""
    d = str(desc or "").upper().replace(" ", "")
    u = float(units or 0)
    if "50LT" in d:
        return u * (50.0 / 7.8)
    if "30LT" in d:
        return u * (30.0 / 7.8)
    if "20LT" in d:
        return u * (20.0 / 7.8)
    try:
        b = float(bpc or 0)
    except (TypeError, ValueError):
        b = 0.0
    return (u / b) if b > 0 else 0.0


def _is_keg(desc: str) -> bool:
    d = str(desc or "").upper().replace(" ", "")
    return ("50LT" in d) or ("30LT" in d) or ("20LT" in d)


@st.cache_data(ttl=3600, show_spinner=False)
def _load_stock_baseline() -> tuple[str, dict]:
    """Per-item closing stock (bottles) from the last ERP Stock & Sale export
    — the authoritative anchor. Returns (baseline_date 'YYYY-MM-DD',
    {ItemID: bottles}). Empty if no baseline file is present."""
    try:
        with open(_BASELINE_PATH, encoding="utf-8") as f:
            d = json.load(f)
        items = {str(k): int(v) for k, v in d.get("items", {}).items()}
        return str(d.get("baseline_date", "")), items
    except Exception as exc:
        print(f"[inventory] stock baseline unavailable: {exc}", file=sys.stderr)
        return "", {}


@st.cache_data(ttl=300, show_spinner=False)
def _load_rollforward(after_date: str) -> dict:
    """Net stock movement per item (bottles, In − Out, FREE GOODS INCLUDED)
    for vouchers dated ON/AFTER the baseline date through today. FY-CASE
    deduped. The baseline is the FY-opening stock (start of 1 April), so we
    include movements from that date onward (>=). Free goods are included
    because dispatched free stock physically leaves the godown (matches the
    ERP Stock & Sale 'Out')."""
    if not after_date:
        return {}
    sql = f"""
        SELECT vi.ItemID,
            SUM(CASE WHEN mt.QtyInOut='I' THEN ISNULL(vi.TotalBottleQty,0)
                     WHEN mt.QtyInOut='O' THEN -ISNULL(vi.TotalBottleQty,0)
                     ELSE 0 END)                          AS NetB
        FROM TrVocItem vi
        JOIN TrVocHead   h  ON h.TransTypeID = vi.TransTypeID AND h.VoucherNo = vi.VoucherNo
        JOIN MsTransType mt ON mt.TransTypeID = vi.TransTypeID
        WHERE h.Cancelled = 'N' AND mt.ItemYN = 'Y' AND mt.QtyInOut IN ('I','O')
          AND vi.ItemID LIKE 'I%'
          AND CAST(h.VoucherDate AS date) >= ?
          AND CAST(h.VoucherDate AS date) <= CAST(GETDATE() AS date)
          {_FY_JOIN}
        GROUP BY vi.ItemID
    """
    df = run_query(sql, (after_date,))
    if df.empty:
        return {}
    return {str(r.ItemID).strip(): int(r.NetB or 0) for r in df.itertuples()}


@st.cache_data(ttl=300, show_spinner=False)
def _load_rollforward_io(after_date: str) -> pd.DataFrame:
    """Per-item Inward / Outward bottles (free goods INCLUDED) on/after the
    baseline date — the split version of _load_rollforward, used by the stock
    reconciliation so Opening + In − Out ties exactly to the live Closing."""
    cols = ["ItemID", "InB", "OutB"]
    if not after_date:
        return pd.DataFrame(columns=cols)
    sql = f"""
        SELECT vi.ItemID,
            SUM(CASE WHEN mt.QtyInOut='I' THEN ISNULL(vi.TotalBottleQty,0) ELSE 0 END) AS InB,
            SUM(CASE WHEN mt.QtyInOut='O' THEN ISNULL(vi.TotalBottleQty,0) ELSE 0 END) AS OutB
        FROM TrVocItem vi
        JOIN TrVocHead   h  ON h.TransTypeID = vi.TransTypeID AND h.VoucherNo = vi.VoucherNo
        JOIN MsTransType mt ON mt.TransTypeID = vi.TransTypeID
        WHERE h.Cancelled = 'N' AND mt.ItemYN = 'Y' AND mt.QtyInOut IN ('I','O')
          AND vi.ItemID LIKE 'I%'
          AND CAST(h.VoucherDate AS date) >= ?
          AND CAST(h.VoucherDate AS date) <= CAST(GETDATE() AS date)
          {_FY_JOIN}
        GROUP BY vi.ItemID
    """
    df = run_query(sql, (after_date,))
    if df.empty:
        return pd.DataFrame(columns=cols)
    df["InB"]  = pd.to_numeric(df["InB"],  errors="coerce").fillna(0).astype(int)
    df["OutB"] = pd.to_numeric(df["OutB"], errors="coerce").fillna(0).astype(int)
    return df


@st.cache_data(ttl=300, show_spinner=False)
def _load_current_stock() -> pd.DataFrame:
    """Live closing stock per item = ERP Stock & Sale baseline + live movements
    since the baseline date. Ties to the ERP report and stays current as bills
    post — no daily export needed. Falls back to the MsItemBatchOpening batch
    table only if no baseline file is present."""
    base_date, base = _load_stock_baseline()
    if not base:
        return _load_current_stock_legacy()
    roll = _load_rollforward(base_date)
    item_ids = set(base) | set(roll)
    ph = ",".join(f"'{i}'" for i in item_ids)
    df = run_query(f"""
        SELECT im.ItemID,
            ISNULL(im.ItemDescription, im.ItemID)        AS ItemDescription,
            ISNULL(b.BrandName,        '(unknown)')      AS BrandName,
            ISNULL(b.CompanyID,        '')               AS CompanyID,
            ISNULL(im.BottlesPerCase,  0)                AS BottlesPerCase,
            ISNULL(im.ValuationCaseRate,   0.0)          AS ValRateCase,
            ISNULL(im.ValuationBottleRate, 0.0)          AS ValRateBottle
        FROM MsItemMaster im
        LEFT JOIN MsBrandMaster b ON b.BrandID = im.BrandID
        WHERE im.ItemID IN ({ph})
    """)
    if df.empty:
        return df
    df["BottlesPerCase"] = pd.to_numeric(df["BottlesPerCase"], errors="coerce").fillna(0).astype(int)
    df["ValRateCase"]    = pd.to_numeric(df["ValRateCase"],    errors="coerce").fillna(0.0)
    df["ValRateBottle"]  = pd.to_numeric(df["ValRateBottle"],  errors="coerce").fillna(0.0)
    # Clamp at 0: re-coded SKUs (price change → new ItemID) can roll forward
    # negative when their pre-recode outflow lands on the new code. A negative
    # physical stock is meaningless and would silently drag down totals.
    df["ClosingBottles"] = df["ItemID"].map(
        lambda i: base.get(i, 0) + roll.get(i, 0)).clip(lower=0).astype(int)
    df["Principal"]      = df["CompanyID"].map(_PRINCIPAL_NAMES).fillna("Other")
    df["ClosingCases"]   = df.apply(
        lambda r: _kegaware_cases(r["ItemDescription"], r["ClosingBottles"],
                                  r["BottlesPerCase"]), axis=1)
    df["ClosingCasesPlain"] = df.apply(
        lambda r: (r["ClosingBottles"] / r["BottlesPerCase"])
                  if r["BottlesPerCase"] > 0 else 0.0, axis=1)
    df["CaseRem"]   = df["ClosingCases"].astype(int)
    df["BottleRem"] = df.apply(
        lambda r: 0 if _is_keg(r["ItemDescription"])
                  else (int(r["ClosingBottles"] - r["CaseRem"] * r["BottlesPerCase"])
                        if r["BottlesPerCase"] > 0 else int(r["ClosingBottles"])),
        axis=1)
    return df


@st.cache_data(ttl=3600, show_spinner=False)
def _load_current_stock_legacy() -> pd.DataFrame:
    """Fallback: live closing stock from MsItemBatchOpening (used only when no
    S&S baseline file exists). NOTE: this batch table doesn't reconcile to the
    ERP Stock & Sale report for re-coded SKUs — that's why the baseline exists."""
    sql = """
        SELECT
            bo.ItemID,
            ISNULL(im.ItemDescription, bo.ItemID)        AS ItemDescription,
            ISNULL(b.BrandName,        '(unknown)')      AS BrandName,
            ISNULL(b.CompanyID,        '')               AS CompanyID,
            ISNULL(im.BottlesPerCase,  0)                AS BottlesPerCase,
            ISNULL(im.ValuationCaseRate,   0.0)          AS ValRateCase,
            ISNULL(im.ValuationBottleRate, 0.0)          AS ValRateBottle,
            SUM(ISNULL(bo.ClosingQty, 0))                AS ClosingBottles
        FROM MsItemBatchOpening bo
        LEFT JOIN MsItemMaster  im ON im.ItemID  = bo.ItemID
        LEFT JOIN MsBrandMaster b  ON b.BrandID  = im.BrandID
        WHERE bo.ItemID LIKE 'I%'
        GROUP BY bo.ItemID, im.ItemDescription, b.BrandName, b.CompanyID,
                 im.BottlesPerCase, im.ValuationCaseRate, im.ValuationBottleRate
    """
    df = run_query(sql)
    if not df.empty:
        df["BottlesPerCase"] = pd.to_numeric(df["BottlesPerCase"], errors="coerce").fillna(0).astype(int)
        df["ClosingBottles"] = pd.to_numeric(df["ClosingBottles"], errors="coerce").fillna(0).astype(int)
        df["ValRateCase"]    = pd.to_numeric(df["ValRateCase"],    errors="coerce").fillna(0.0)
        df["ValRateBottle"]  = pd.to_numeric(df["ValRateBottle"],  errors="coerce").fillna(0.0)
        df["Principal"]      = df["CompanyID"].map(_PRINCIPAL_NAMES).fillna("Other")
        df["ClosingCases"]   = df.apply(
            lambda r: _kegaware_cases(r["ItemDescription"], r["ClosingBottles"],
                                      r["BottlesPerCase"]),
            axis=1,
        )
        df["ClosingCasesPlain"] = df.apply(
            lambda r: (r["ClosingBottles"] / r["BottlesPerCase"])
                      if r["BottlesPerCase"] > 0 else 0.0,
            axis=1,
        )
        df["CaseRem"]      = df["ClosingCases"].astype(int)
        df["BottleRem"]    = df.apply(
            lambda r: 0 if _is_keg(r["ItemDescription"])
                      else (int(r["ClosingBottles"] - r["CaseRem"] * r["BottlesPerCase"])
                            if r["BottlesPerCase"] > 0 else int(r["ClosingBottles"])),
            axis=1,
        )
    return df


@st.cache_data(ttl=3600, show_spinner=False)
def _load_movements(start: date, end: date, keg_aware: bool = True) -> pd.DataFrame:
    """Per-item In/Out bottle + case movements between start and end.

    Uses the FY-CASE filter so duplicate-FY rows in TrVocItem are dropped.
    Verified vs ERP Stock-Balance report (FY 2025-26): In within 0.005%.
    keg_aware toggles volume-conversion of kegs in the case columns.
    """
    _CX = cases_sql(keg_aware)
    sql = f"""
        SELECT
            vi.ItemID,
            SUM(CASE WHEN mt.QtyInOut='I' THEN ISNULL(vi.TotalBottleQty,0) ELSE 0 END) AS InBottles,
            SUM(CASE WHEN mt.QtyInOut='O' THEN ISNULL(vi.TotalBottleQty,0) ELSE 0 END) AS OutBottles,
            SUM(CASE WHEN mt.QtyInOut='I' THEN {_CX} ELSE 0 END) AS InCases,
            SUM(CASE WHEN mt.QtyInOut='O' THEN {_CX} ELSE 0 END) AS OutCases
        FROM TrVocItem vi
        JOIN TrVocHead   h  ON h.TransTypeID = vi.TransTypeID AND h.VoucherNo = vi.VoucherNo
        JOIN MsTransType mt ON mt.TransTypeID = vi.TransTypeID
        JOIN MsItemMaster im ON im.ItemID    = vi.ItemID
        WHERE h.Cancelled  = 'N'
          AND vi.FreeItemYN = 'N'
          AND vi.ItemID     LIKE 'I%'
          AND mt.ItemYN     = 'Y'
          AND mt.QtyInOut IN ('I','O')
          AND h.VoucherDate BETWEEN ? AND ?
          {_FY_JOIN}
        GROUP BY vi.ItemID
    """
    df = run_query(sql, (str(start), str(end)))
    if not df.empty:
        for c in ("InBottles", "OutBottles"):
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)
        for c in ("InCases", "OutCases"):
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    return df


@st.cache_data(ttl=86400, show_spinner=False)   # FY-fixed — 24h
def _load_opening_stock() -> pd.DataFrame:
    """Per-item opening bottles from MsItemBatchOpening (FY-opening basis)."""
    sql = """
        SELECT ItemID,
               SUM(ISNULL(OpeningQty, 0))  AS OpeningBottles
        FROM MsItemBatchOpening
        WHERE ItemID LIKE 'I%'
        GROUP BY ItemID
    """
    df = run_query(sql)
    if not df.empty:
        df["OpeningBottles"] = pd.to_numeric(df["OpeningBottles"], errors="coerce").fillna(0).astype(int)
    return df


@st.cache_data(ttl=86400, show_spinner=False)   # item classification — 24h
def _load_item_origin() -> pd.DataFrame:
    """Classify each item as Import / Daman / Domestic based on dominant
    purchase TransType in the last 12 months."""
    imp_ph    = ",".join(str(t) for t in IMPORT_TYPES)
    daman_ph  = ",".join(str(t) for t in DAMAN_TYPES)
    pu_ph     = ",".join(str(t) for t in PURCHASE_TYPES)
    sql = f"""
        WITH PerTT AS (
            SELECT
                vi.ItemID, h.TransTypeID,
                SUM(ISNULL(vi.TotalBottleQty, 0)) AS Bottles
            FROM TrVocItem vi
            JOIN TrVocHead h ON h.TransTypeID = vi.TransTypeID AND h.VoucherNo = vi.VoucherNo
            WHERE h.Cancelled  = 'N'
              AND vi.FreeItemYN = 'N'
              AND vi.ItemID     LIKE 'I%'
              AND h.TransTypeID IN ({pu_ph})
              AND h.VoucherDate >= DATEADD(MONTH, -12, GETDATE())
              {_FY_JOIN}
            GROUP BY vi.ItemID, h.TransTypeID
        )
        SELECT
            ItemID,
            CASE
              WHEN MAX(CASE WHEN TransTypeID IN ({imp_ph})   THEN Bottles ELSE 0 END) > 0 THEN 'Import'
              WHEN MAX(CASE WHEN TransTypeID IN ({daman_ph}) THEN Bottles ELSE 0 END) > 0 THEN 'Daman'
              ELSE 'Domestic'
            END AS Origin
        FROM PerTT
        GROUP BY ItemID
    """
    return run_query(sql)


@st.cache_data(ttl=3600, show_spinner=False)
def _load_sales_velocity(as_of_date: date, days_back: int = 30,
                         keg_aware: bool = True) -> pd.DataFrame:
    """Per-item sales cases in last N days + last sale date (FY-CASE filtered)."""
    sql = f"""
        SELECT
            vi.ItemID,
            SUM({cases_sql(keg_aware)})               AS Cases,
            MAX(h.VoucherDate)                        AS LastSale
        FROM TrVocItem vi
        JOIN TrVocHead    h  ON h.TransTypeID = vi.TransTypeID AND h.VoucherNo = vi.VoucherNo
        JOIN MsTransType  mt ON mt.TransTypeID = vi.TransTypeID
        JOIN MsItemMaster im ON im.ItemID      = vi.ItemID
        WHERE h.Cancelled  = 'N'
          AND mt.QtyInOut  = 'O'
          AND vi.FreeItemYN = 'N'
          AND vi.ItemID     LIKE 'I%'
          AND h.VoucherDate BETWEEN DATEADD(DAY, -{days_back}, ?) AND ?
          {_FY_JOIN}
        GROUP BY vi.ItemID
    """
    df = run_query(sql, (str(as_of_date), str(as_of_date)))
    if not df.empty:
        df["Cases"]    = pd.to_numeric(df["Cases"], errors="coerce").fillna(0.0)
        df["LastSale"] = pd.to_datetime(df["LastSale"], errors="coerce")
    return df


@st.cache_data(ttl=600, show_spinner=False)
def _load_demand_signals(as_of: date, keg_aware: bool = True) -> pd.DataFrame:
    """Per-item SALES demand signals for the indent predictor:
    L3M = avg/month over the 3 prior full months · PrevMo = last full month ·
    ThisMTD = sold so far this month. SALES_TYPES only (true outlet demand)."""
    msf   = _month_first(as_of)               # 1st of this month
    l3s   = _month_first(as_of, -3)           # 1st, 3 months back
    prevs = _month_first(as_of, -1)           # 1st of last month
    preve = msf - timedelta(days=1)           # last day of last month
    sales = ",".join(str(t) for t in SALES_TYPES)
    cx = cases_sql(keg_aware)
    sql = f"""
        SELECT vi.ItemID,
            SUM(CASE WHEN h.VoucherDate >= ? AND h.VoucherDate < ?  THEN {cx} ELSE 0 END) AS L3MCases,
            SUM(CASE WHEN h.VoucherDate >= ? AND h.VoucherDate <= ? THEN {cx} ELSE 0 END) AS PrevMo,
            SUM(CASE WHEN h.VoucherDate >= ? AND h.VoucherDate <= ? THEN {cx} ELSE 0 END) AS ThisMTD
        FROM TrVocItem vi
        JOIN TrVocHead    h  ON h.TransTypeID = vi.TransTypeID AND h.VoucherNo = vi.VoucherNo
        JOIN MsItemMaster im ON im.ItemID      = vi.ItemID
        WHERE h.Cancelled = 'N' AND vi.FreeItemYN = 'N' AND vi.ItemID LIKE 'I%'
          AND h.TransTypeID IN ({sales})
          AND h.VoucherDate >= ? AND h.VoucherDate <= ?
          {_FY_JOIN}
        GROUP BY vi.ItemID
    """
    p = (str(l3s), str(msf), str(prevs), str(preve), str(msf), str(as_of),
         str(l3s), str(as_of))
    df = run_query(sql, p)
    if df.empty:
        return pd.DataFrame(columns=["ItemID", "L3MMonthly", "PrevMo", "ThisMTD"])
    for c in ("L3MCases", "PrevMo", "ThisMTD"):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
    df["L3MMonthly"] = df["L3MCases"] / 3.0
    return df[["ItemID", "L3MMonthly", "PrevMo", "ThisMTD"]]


# ═══════════════════════════════════════════════════════════════════════════════
# COMPUTATION HELPERS
# ═══════════════════════════════════════════════════════════════════════════════

def _kpi_card(label: str, value: str, sub: str,
              sub_color: str, accent: str) -> str:
    return f"""
    <div style='background:#fff;border:1px solid #e5e7eb;border-radius:8px;
                border-left:4px solid {accent};padding:12px 16px;
                box-shadow:0 1px 2px rgba(0,0,0,0.04);'>
        <div style='font-size:0.7rem;color:#6b7280;
                    text-transform:uppercase;letter-spacing:0.05em'>{label}</div>
        <div style='font-size:1.5rem;font-weight:700;color:#111827;
                    margin-top:4px;line-height:1.15'>{value}</div>
        <div style='font-size:0.78rem;color:{sub_color};margin-top:2px;
                    font-weight:600'>{sub}</div>
    </div>
    """


def _fmt_cases_with_remainder(cases_int: int, bottle_rem: int) -> str:
    if bottle_rem > 0:
        return f"{cases_int:,} cs + {bottle_rem:,} bot"
    return f"{cases_int:,} cs"


def _build_stock_df(as_of: date) -> pd.DataFrame:
    """Stock per item as of `as_of`. Live for today; rolled back otherwise.

    Live source: MsItemBatchOpening.ClosingQty
    Historical:  Live - movements after as_of (now using FY-CASE filter)
    """
    stock_df = _load_current_stock()
    if stock_df.empty:
        return stock_df

    today = date.today()
    if as_of >= today:
        out = stock_df.copy()
    else:
        # Movements strictly AFTER as_of through today, FY-CASE filtered
        moves_after = _load_movements(as_of, today)
        if moves_after.empty:
            out = stock_df.copy()
        else:
            merged = stock_df.merge(
                moves_after[["ItemID", "InBottles", "OutBottles"]],
                on="ItemID", how="left",
            ).fillna({"InBottles": 0, "OutBottles": 0})
            merged["ClosingBottles"] = (
                merged["ClosingBottles"] + merged["OutBottles"] - merged["InBottles"]
            ).clip(lower=0).astype(int)
            out = merged.drop(columns=["InBottles", "OutBottles"], errors="ignore")

    # Recompute cases/remainder from updated ClosingBottles (keg-aware)
    out["ClosingCases"] = out.apply(
        lambda r: _kegaware_cases(r["ItemDescription"], r["ClosingBottles"],
                                  r["BottlesPerCase"]),
        axis=1,
    )
    out["ClosingCasesPlain"] = out.apply(
        lambda r: (r["ClosingBottles"] / r["BottlesPerCase"])
                  if r["BottlesPerCase"] > 0 else 0.0,
        axis=1,
    )
    out["CaseRem"]   = out["ClosingCases"].astype(int)
    out["BottleRem"] = out.apply(
        lambda r: 0 if _is_keg(r["ItemDescription"])
                  else (int(r["ClosingBottles"] - r["CaseRem"] * r["BottlesPerCase"])
                        if r["BottlesPerCase"] > 0 else int(r["ClosingBottles"])),
        axis=1,
    )
    return out


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION RENDERERS
# ═══════════════════════════════════════════════════════════════════════════════

def _section_kpis(stock_df: pd.DataFrame) -> None:
    in_stock = stock_df[stock_df["ClosingBottles"] > 0]
    total_items   = len(in_stock)
    total_bottles = int(in_stock["ClosingBottles"].sum())
    total_cases   = float(in_stock["ClosingCases"].sum())

    # Inventory value uses Valuation rate (already includes excise / landed cost)
    merged = stock_df.copy()
    merged["Value"] = merged["ClosingCasesPlain"] * merged["ValRateCase"]
    total_value = float(merged["Value"].sum())

    out_of_stock = int((stock_df["ClosingBottles"] <= 0).sum())

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(_kpi_card(
            "Items in Stock",
            f"{total_items:,}",
            f"of {len(stock_df):,} total items",
            "#6b7280", _KPI_COLORS[0],
        ), unsafe_allow_html=True)
    with c2:
        # Cases first, bottles secondary
        cs_int = int(total_cases)
        bot_rem = total_bottles - cs_int * (total_bottles // cs_int if cs_int else 0)
        bot_only = total_bottles  # show all bottles in sub
        st.markdown(_kpi_card(
            "Total Cases",
            f"{cs_int:,}",
            f"{total_bottles:,} bottles total",
            "#6b7280", _KPI_COLORS[1],
        ), unsafe_allow_html=True)
    with c3:
        st.markdown(_kpi_card(
            "Estimated Value",
            f"₹{total_value/1e7:.2f} Cr",
            "At MsItemMaster ValuationCaseRate",
            "#6b7280", _KPI_COLORS[2],
        ), unsafe_allow_html=True)
    with c4:
        st.markdown(_kpi_card(
            "Out of Stock",
            f"{out_of_stock:,}",
            "items with zero closing",
            "#dc2626" if out_of_stock else "#6b7280", _KPI_COLORS[3],
        ), unsafe_allow_html=True)


def _section_reconciliation(stock_df: pd.DataFrame,
                            keg_aware: bool = True) -> None:
    """Stock flow this FY: Opening (1 Apr) + Inward − Outward = Closing (now).
    Sourced consistently with the live stock — Opening from the S&S baseline,
    In/Out from live movements since — so the four numbers always tie out."""
    st.markdown("##### Stock flow this year")

    base_date, base = _load_stock_baseline()
    io = _load_rollforward_io(base_date)

    # One row per item: opening + inward − outward = closing (bottles)
    meta = stock_df[["ItemID", "BottlesPerCase", "ItemDescription"]].drop_duplicates("ItemID")
    m = meta.copy()
    m["OpenB"] = m["ItemID"].map(lambda i: base.get(i, 0)).fillna(0)
    io_in  = dict(zip(io["ItemID"], io["InB"]))  if not io.empty else {}
    io_out = dict(zip(io["ItemID"], io["OutB"])) if not io.empty else {}
    m["InB"]  = m["ItemID"].map(lambda i: io_in.get(i, 0))
    m["OutB"] = m["ItemID"].map(lambda i: io_out.get(i, 0))
    m["CloseB"] = m["OpenB"] + m["InB"] - m["OutB"]
    m["BottlesPerCase"]  = pd.to_numeric(m["BottlesPerCase"], errors="coerce").fillna(0).astype(int)
    m["ItemDescription"] = m["ItemDescription"].fillna("")

    def _cs(col):
        return m.apply(lambda r: _kegaware_cases(r["ItemDescription"], r[col],
                                                 r["BottlesPerCase"]) if keg_aware
                       else (r[col] / r["BottlesPerCase"] if r["BottlesPerCase"] > 0 else 0.0),
                       axis=1).sum()

    opening, inward, outward, closing = _cs("OpenB"), _cs("InB"), _cs("OutB"), _cs("CloseB")

    bd = base_date or "FY start"
    st.caption(
        f"**Opening stock ({bd})  +  Inward  −  Outward  =  Closing stock (today).** "
        "Opening is the ERP Stock & Sale year-opening; In/Out are live bill "
        "movements (free goods included). Cases shown rounded."
    )
    g1, g2, g3, g4 = st.columns(4)
    g1.metric("Opening (1 Apr)", f"{opening:,.0f} cs")
    g2.metric("➕ Inward",        f"{inward:,.0f} cs")
    g3.metric("➖ Outward",       f"{outward:,.0f} cs")
    g4.metric("🟰 Closing (now)", f"{closing:,.0f} cs")


def _section_by_principal(stock_df: pd.DataFrame) -> None:
    st.markdown("##### Stock by Principal")
    if stock_df.empty:
        st.info("No stock data."); return

    df = stock_df.copy()
    df["Value"] = df["ClosingCasesPlain"] * df["ValRateCase"]
    g = (
        df[df["ClosingBottles"] > 0]
        .groupby("Principal", as_index=False)
        .agg(Items=("ItemID", "nunique"),
             Cases=("ClosingCases", "sum"),
             Value=("Value", "sum"))
        .sort_values("Value", ascending=False)
    )
    st.dataframe(
        g.rename(columns={
            "Items": "Items in Stock", "Cases": "Total Cases",
            "Value": "Estimated Value ₹",
        }).style.format({
            "Items in Stock":     "{:,}",
            "Total Cases":        "{:,.0f}",
            "Estimated Value ₹":  format_inr,
        }),
        use_container_width=True, hide_index=True,
    )


def _section_top_items(stock_df: pd.DataFrame, origin_df: pd.DataFrame) -> None:
    st.markdown("##### Top 20 items by stock value")
    if stock_df.empty:
        st.info("No stock data."); return

    df = stock_df.merge(origin_df, on="ItemID", how="left").fillna({"Origin": "Domestic"})
    df["Value"] = df["ClosingCasesPlain"] * df["ValRateCase"]
    top = df[df["Value"] > 0].sort_values("Value", ascending=False).head(20).copy()
    if top.empty:
        st.info("No items with value > 0."); return

    # Bar chart by principal color
    top_chart = top.sort_values("Value", ascending=True)
    colors = [_PRINCIPAL_COLOR.get(p, "#B4B2A9") for p in top_chart["Principal"]]
    top_chart["ValueCr"] = top_chart["Value"] / 1e7
    fig = go.Figure(go.Bar(
        x=top_chart["ValueCr"], y=top_chart["ItemDescription"],
        orientation="h", marker_color=colors,
        text=[f"₹{v:.2f} Cr" for v in top_chart["ValueCr"]],
        textposition="outside",
        customdata=top_chart["Value"].apply(format_inr),
        hovertemplate="<b>%{y}</b><br>%{customdata}<extra></extra>",
    ))
    fig.update_layout(
        paper_bgcolor="rgba(0,0,0,0)", plot_bgcolor="rgba(0,0,0,0)",
        template="plotly_white",
        margin=dict(l=16, r=16, t=16, b=16),
        height=max(420, len(top_chart) * 24),
        xaxis=dict(title="Value (₹ Cr)", ticksuffix=" Cr", gridcolor="#E8E8E8"),
        yaxis=dict(gridcolor="#E8E8E8"),
    )
    st.plotly_chart(fig, use_container_width=True)

    # Companion table with Origin column + duplicate items kept separate
    top["Stock"] = top.apply(
        lambda r: _fmt_cases_with_remainder(int(r["CaseRem"]), int(r["BottleRem"])),
        axis=1,
    )
    tbl = top[["BrandName", "ItemDescription", "Origin",
               "Stock", "ValRateCase", "Value"]].rename(columns={
        "BrandName":       "Brand",
        "ItemDescription": "Item",
        "ValRateCase":     "Landed Rate",
        "Value":           "Stock Value",
    })
    st.dataframe(
        tbl.style.format({
            "Landed Rate":  "₹{:,.0f}",
            "Stock Value":  format_inr,
        }),
        use_container_width=True, hide_index=True,
    )


def _section_slow_movers(stock_df: pd.DataFrame, vel_df: pd.DataFrame,
                         threshold_cases: int = 50) -> None:
    st.markdown("##### Slow Movers — Capital tied up")
    st.caption(f"Closing > {threshold_cases} cases AND last-30d sales < 10 cases.")
    if stock_df.empty:
        st.info("No stock data."); return

    merged = stock_df.merge(vel_df[["ItemID", "Cases", "LastSale"]],
                            on="ItemID", how="left").rename(columns={"Cases": "Sales30"})
    merged["Sales30"] = pd.to_numeric(merged["Sales30"], errors="coerce").fillna(0.0)
    # Valuation is ALWAYS on physical (un-converted) units — a keg is one keg,
    # never 6.41 cases. Keg-aware cases are only for purchase/sales reporting.
    merged["Value"]   = merged["ClosingCasesPlain"] * merged["ValRateCase"]

    slow = merged[
        (merged["ClosingCases"] > threshold_cases) & (merged["Sales30"] < 10)
    ].copy()
    if slow.empty:
        st.success("No slow movers."); return

    slow["DailyAvg"]  = slow["Sales30"] / 30
    slow["DaysCover"] = slow.apply(
        lambda r: (r["ClosingCases"] / r["DailyAvg"]) if r["DailyAvg"] > 0 else 9999,
        axis=1,
    )
    slow = slow.sort_values("DaysCover", ascending=False).head(30)

    # Keep original column names so styler + format always agree.
    # Display labels are handled via Streamlit column_config below.
    disp = slow[["BrandName", "ItemDescription", "ClosingCases",
                 "Sales30", "DaysCover", "Value"]].copy()

    def _row_style(row):
        v = row["DaysCover"]
        if v > 90:    bg = "background-color:#fee2e2"
        elif v > 30:  bg = "background-color:#fef3c7"
        else:         bg = ""
        return [bg] * len(row)

    fmt = {
        "ClosingCases": "{:,.0f}",
        "Sales30":      "{:,.2f}",
        "DaysCover":    lambda x: "9999+" if x >= 9999 else f"{x:,.0f}",
        "Value":        format_inr,
    }
    styled = (
        disp.style
        .apply(_row_style, axis=1)
        .format(_safe_format(disp, fmt))
    )
    st.dataframe(
        styled,
        use_container_width=True,
        hide_index=True,
        column_config={
            "BrandName":       st.column_config.Column("Brand"),
            "ItemDescription": st.column_config.Column("Item"),
            "ClosingCases":    st.column_config.Column("Closing Cases"),
            "Sales30":         st.column_config.Column("Last 30d Cases"),
            "DaysCover":       st.column_config.Column("Days of Cover"),
            "Value":           st.column_config.Column("Stock Value"),
        },
    )


def _section_out_of_stock(stock_df: pd.DataFrame, vel_30: pd.DataFrame,
                          vel_90: pd.DataFrame) -> None:
    st.markdown("##### ⚠️ Out of Stock — Risk of lost sales")
    st.caption("Items with zero closing AND proven demand (>5 cases in last 30 days).")
    if stock_df.empty:
        st.info("No stock data."); return

    s = stock_df.merge(vel_30[["ItemID", "Cases", "LastSale"]],
                       on="ItemID", how="left").rename(columns={"Cases": "Sales30"})
    s["Sales30"] = pd.to_numeric(s["Sales30"], errors="coerce").fillna(0.0)
    s = s.merge(vel_90[["ItemID", "Cases"]].rename(columns={"Cases": "Sales90"}),
                on="ItemID", how="left")
    s["Sales90"] = pd.to_numeric(s["Sales90"], errors="coerce").fillna(0.0)

    risk = s[(s["ClosingCases"] <= 0) & (s["Sales30"] > 5)].copy()
    if risk.empty:
        st.success("No out-of-stock items with recent demand."); return

    risk = risk.sort_values("Sales30", ascending=False).head(40)
    risk["Last Sale"] = pd.to_datetime(risk["LastSale"], errors="coerce") \
        .dt.strftime("%d %b %Y").fillna("—")
    disp = risk[["BrandName", "ItemDescription", "Last Sale",
                 "Sales30", "Sales90"]].rename(columns={
        "BrandName":       "Brand",
        "ItemDescription": "Item",
        "Sales30":         "Last 30d Cases",
        "Sales90":         "Last 90d Cases",
    })
    st.dataframe(
        disp.style.format({"Last 30d Cases": "{:,.2f}", "Last 90d Cases": "{:,.2f}"}),
        use_container_width=True, hide_index=True,
    )


def _section_days_of_cover(stock_df: pd.DataFrame, vel_df: pd.DataFrame) -> None:
    st.markdown("##### Days of Cover — Top 30 selling items")
    st.caption("Green > 30 days · Amber 15–30 · Red < 15 (urgent reorder)")
    if stock_df.empty or vel_df.empty:
        st.info("Insufficient data."); return

    merged = stock_df.merge(vel_df[["ItemID", "Cases"]],
                            on="ItemID", how="left").rename(columns={"Cases": "Sales30"})
    merged["Sales30"] = pd.to_numeric(merged["Sales30"], errors="coerce").fillna(0.0)
    sellers = merged[merged["Sales30"] > 0].copy()
    if sellers.empty:
        st.info("No selling items in the last 30 days."); return
    sellers["DailyAvg"]  = sellers["Sales30"] / 30
    sellers["DaysCover"] = sellers.apply(
        lambda r: (r["ClosingCases"] / r["DailyAvg"]) if r["DailyAvg"] > 0 else 9999,
        axis=1,
    )
    top_sellers = sellers.sort_values("Sales30", ascending=False).head(30)

    # Style on original column names; show display labels via column_config.
    disp = top_sellers[["BrandName", "ItemDescription", "ClosingCases",
                        "DailyAvg", "DaysCover"]].copy()

    def _doc_style(row):
        v = row["DaysCover"]
        if v >= 9999:        bg = ""
        elif v >= 30:        bg = "background-color:#dcfce7"
        elif v >= 15:        bg = "background-color:#fef3c7"
        else:                bg = "background-color:#fee2e2"
        return [bg if c == "DaysCover" else "" for c in row.index]

    fmt = {
        "ClosingCases": "{:,.0f}",
        "DailyAvg":     "{:.2f}",
        "DaysCover":    lambda x: "9999+" if x >= 9999 else f"{x:,.0f}",
    }
    styled = (
        disp.style
        .apply(_doc_style, axis=1)
        .format(_safe_format(disp, fmt))
    )
    st.dataframe(
        styled,
        use_container_width=True,
        hide_index=True,
        column_config={
            "BrandName":       st.column_config.Column("Brand"),
            "ItemDescription": st.column_config.Column("Item"),
            "ClosingCases":    st.column_config.Column("Stock (Cases)"),
            "DailyAvg":        st.column_config.Column("Daily Avg Sales"),
            "DaysCover":       st.column_config.Column("Days of Cover"),
        },
    )


def _section_indent_planner(stock_df: pd.DataFrame, as_of: date,
                            keg_aware: bool = True) -> None:
    """Predict the stock you'll fall short of this month, so you can indent
    early. Per SKU: projected demand = balanced blend of L3M monthly, last
    month, and this-month run-rate; remaining = projected − sold MTD; if
    remaining > stock on hand, the gap is the indent quantity."""
    st.markdown("##### 📋 Indent planner — stock you may fall short of")
    days_in_month = calendar.monthrange(as_of.year, as_of.month)[1]
    days_done = max(as_of.day, 1)
    st.caption(
        f"Projected demand = balanced blend of L3M monthly avg, last month, and "
        f"this-month run-rate (day {days_done}/{days_in_month}). **Indent** = "
        f"demand still expected this month − stock on hand. Order these early."
    )

    dem = _load_demand_signals(as_of, keg_aware=keg_aware)
    if dem.empty:
        st.info("No recent sales to forecast from."); return
    m = stock_df.merge(dem, on="ItemID", how="left")
    for c in ("L3MMonthly", "PrevMo", "ThisMTD"):
        m[c] = pd.to_numeric(m.get(c, 0), errors="coerce").fillna(0.0)

    # Run-rate projection of this month to a full month
    m["ThisProj"] = m["ThisMTD"] / days_done * days_in_month
    # Balanced blend (35% L3M / 30% last month / 35% this-month run-rate)
    m["Projected"] = 0.35 * m["L3MMonthly"] + 0.30 * m["PrevMo"] + 0.35 * m["ThisProj"]
    m["Remaining"] = (m["Projected"] - m["ThisMTD"]).clip(lower=0)
    m["Indent"]    = (m["Remaining"] - m["ClosingCases"]).clip(lower=0)
    daily = (m["L3MMonthly"] / 30.0).replace(0, pd.NA)
    m["DaysCover"] = (m["ClosingCases"] / daily).fillna(9999)

    short = m[m["Indent"] >= 1].sort_values("Indent", ascending=False)
    if short.empty:
        st.success("✅ Enough stock on hand to meet this month's projected demand "
                   "for every SKU — nothing to indent right now.")
        return

    st.warning(f"⚠️ **{len(short)} SKUs** projected to fall short this month — "
               f"total indent **{short['Indent'].sum():,.0f} cases**.")
    disp = short[["BrandName", "ItemDescription", "ClosingCases", "L3MMonthly",
                  "PrevMo", "ThisMTD", "Projected", "Remaining", "Indent",
                  "DaysCover"]].copy()
    fmt = {c: "{:,.0f}" for c in ["ClosingCases", "L3MMonthly", "PrevMo",
                                  "ThisMTD", "Projected", "Remaining", "Indent"]}
    fmt["DaysCover"] = lambda v: "—" if v >= 9999 else f"{v:,.0f} d"
    st.dataframe(
        disp.style.format(_safe_format(disp, fmt)),
        use_container_width=True, hide_index=True, height=440,
        column_config={
            "BrandName":       st.column_config.Column("Brand"),
            "ItemDescription": st.column_config.Column("Item"),
            "ClosingCases":    st.column_config.Column("Stock now"),
            "L3MMonthly":      st.column_config.Column("L3M/mo"),
            "PrevMo":          st.column_config.Column("Last month"),
            "ThisMTD":         st.column_config.Column("Sold MTD"),
            "Projected":       st.column_config.Column("Projected demand"),
            "Remaining":       st.column_config.Column("Still needed"),
            "Indent":          st.column_config.Column("➡️ INDENT cs"),
            "DaysCover":       st.column_config.Column("Days cover"),
        },
    )
    st.download_button(
        "⬇️ Download indent list",
        short[["BrandName", "ItemDescription", "ClosingCases", "Projected",
               "Remaining", "Indent"]].to_csv(index=False).encode("utf-8-sig"),
        file_name=f"indent_planner_{as_of:%Y%m%d}.csv", mime="text/csv",
        key="inv_indent_dl")


# ─────────────────────────────────────────────────────────────────────────────
# Primary Plan / Indent — three modes:
#   1. Analyze company's plan (Excel upload) → per-brand verdict vs stock & L3M
#   2. Build our plan (manual target) → distribute target across brands
#   3. Brindco annual transition (preset) → 12-month phasing for 6 brands
# ─────────────────────────────────────────────────────────────────────────────

_PRINCIPAL_CID_MAP = {
    "United Spirits":   "C00025",
    "United Breweries": "C00039",
    "Diageo":           "C00040",
    "Brown-Forman":     "C00056",
}

# Diageo primary-order codes → MsBrandMaster BrandName substring patterns.
# Curated from the live brand master. Each value is a list of UPPER-cased
# substrings; a brand matches if ANY appears in its BrandName.
DIAGEO_CODE_MAP: dict[str, list[str]] = {
    "BAIL":    ["BAILEYS IRISH"],
    "BAILCS":  ["BAILEYS SALTED"],
    "BAILSC":  ["BAILEYS STRAWB"],
    "CAOL":    ["CAOL"],
    "CARD":    ["CARDHU"],
    "CDS":     ["CARDHU DYN", "CARDHU DARK"],
    "CIRV":    ["CIROC"],
    "CLYN":    ["CLYNELISH"],
    "CRAG":    ["CRAGGANMORE"],
    "DALW":    ["DALWHINNIE"],
    "DONA":    ["DON JULIO ANEJO"],
    "DONB":    ["DON JULIO BLANCO"],
    "DONJ":    ["DON JULIO 1942"],
    "DORP":    ["DON JULIO REPOSADO"],
    "GLEN":    ["GLENKINCHIE"],
    "GODWFS":  ["FRUIT& SPICE", "FRUIT & SPICE"],
    "GODWRR":  ["RICH & ROUND"],
    "GORD":    ["GORDON"],
    "J&B":     ["J & B", "J&B"],
    "JW18S":   ["JOHNNIE WALKER AGED 18", "JOHNNIE WALKER 18"],
    "JWB":     ["JOHNNIE WALKER BLUE"],
    "JWBL":    ["JOHNNIE WALKER BLACK LABEL", "JW. BLACK LABEL"],
    "JWBLGCC": ["JOHNNIE WALKER BLACK LABEL"],
    "JWBLLO":  ["LOWLAND ORIGIN"],
    "JWBLON":  ["BLACK SPEYSIDE ORIGIN"],
    "JWDB":    ["DOUBLE BLACK"],
    "JWGL":    ["JOHNNIE WALKER GOLD"],
    "JWGR":    ["GREEN LABE"],
    "JWRL":    ["JOHNNIE WALKER RED LABEL"],
    "JWRRS":   ["RED RYE"],
    "JWXR":    ["XR21", "JW. & SONS XR"],
    "KOV":     ["KETEL ONE"],
    "LAGA":    ["LAGAVULIN"],
    "OBAN":    ["OBAN"],
    "ROECOW":  ["ROE & CO"],
    "SID12S":  ["GLEN DULLAN 12"],
    "SID15S":  ["GLEN DULLAN 15"],
    "SID18S":  ["GLEN DULLAN 18"],
    "SIN12S":  ["FRUITY DECADENCE"],
    "SMVR":    ["SMIRNOFF TRIPLE DISTILLED"],
    "TA10S":   ["TALISKER"],
    "TANQ":    ["TANQUERAY LONDON"],
    "TANQM":   ["TANQUERAY MALACCA"],
    "TANQR":   ["TANQUERAY RANGPUR"],
    "TANT":    ["TANQUERAY TEN"],
}


@st.cache_data(ttl=1800, show_spinner=False)
def _load_principal_brand_history(company_id: str,
                                  months_back: int = 13) -> pd.DataFrame:
    """Per (BrandName, yyyy-MM) cases sold for ONE principal — last N months.
    Plain cases (kegs as 1) — appropriate for spirits; UBL still works since
    BottlesPerCase is set for keg SKUs to '1', giving a per-keg count."""
    type_ph = ",".join(str(t) for t in SALES_TYPES)
    df = run_query(f"""
        SELECT b.BrandName,
               FORMAT(h.VoucherDate, 'yyyy-MM') AS Mon,
               CAST(SUM(CAST(vi.TotalBottleQty AS decimal(18,4))
                        / NULLIF(im.BottlesPerCase, 0)) AS float) AS Cases
        FROM TrVocHead h
        JOIN TrVocItem vi
            ON  vi.TransTypeID = h.TransTypeID AND vi.VoucherNo = h.VoucherNo
            AND vi.ItemID LIKE 'I%'
            AND vi.FinancialYear = CASE
                WHEN MONTH(h.VoucherDate) >= 4
                THEN CAST(YEAR(h.VoucherDate) AS VARCHAR)+'-'+CAST(YEAR(h.VoucherDate)+1 AS VARCHAR)
                ELSE CAST(YEAR(h.VoucherDate)-1 AS VARCHAR)+'-'+CAST(YEAR(h.VoucherDate) AS VARCHAR)
              END
        JOIN MsItemMaster  im ON im.ItemID = vi.ItemID
        JOIN MsBrandMaster b  ON b.BrandID = im.BrandID
        WHERE h.TransTypeID IN ({type_ph}) AND h.Cancelled = 'N'
          AND b.CompanyID = ?
          AND h.VoucherDate >= DATEADD(MONTH, -{months_back}, GETDATE())
        GROUP BY b.BrandName, FORMAT(h.VoucherDate, 'yyyy-MM')
    """, (company_id,))
    if not df.empty:
        df["Cases"] = pd.to_numeric(df["Cases"], errors="coerce").fillna(0.0)
    return df


@st.cache_data(ttl=900, show_spinner=False)
def _load_principal_received_mtd(company_id: str,
                                 as_of: date) -> pd.DataFrame:
    """Cases RECEIVED (purchase vouchers) per (ItemID, BrandName) from
    the 1st of the as_of month up to as_of. Used in the Build planner so
    "balance to order" reflects only what's still pending from the
    company after subtracting what's already arrived this month."""
    type_ph = ",".join(str(t) for t in PURCHASE_TYPES)
    start = as_of.replace(day=1)
    df = run_query(f"""
        SELECT im.ItemID,
               ISNULL(im.ItemDescription, im.ItemID) AS ItemDescription,
               b.BrandName,
               CAST(SUM(CAST(vi.TotalBottleQty AS decimal(18,4))
                        / NULLIF(im.BottlesPerCase, 0)) AS float) AS Cases
        FROM TrVocHead h
        JOIN TrVocItem vi
            ON  vi.TransTypeID = h.TransTypeID AND vi.VoucherNo = h.VoucherNo
            AND vi.ItemID LIKE 'I%'
            AND vi.FinancialYear = CASE
                WHEN MONTH(h.VoucherDate) >= 4
                THEN CAST(YEAR(h.VoucherDate) AS VARCHAR)+'-'+CAST(YEAR(h.VoucherDate)+1 AS VARCHAR)
                ELSE CAST(YEAR(h.VoucherDate)-1 AS VARCHAR)+'-'+CAST(YEAR(h.VoucherDate) AS VARCHAR)
              END
        JOIN MsItemMaster  im ON im.ItemID = vi.ItemID
        JOIN MsBrandMaster b  ON b.BrandID = im.BrandID
        WHERE h.TransTypeID IN ({type_ph}) AND h.Cancelled = 'N'
          AND b.CompanyID = ?
          AND h.VoucherDate >= ? AND h.VoucherDate <= ?
        GROUP BY im.ItemID, im.ItemDescription, b.BrandName
    """, (company_id, start.isoformat(), as_of.isoformat()))
    if not df.empty:
        df["Cases"] = pd.to_numeric(df["Cases"], errors="coerce").fillna(0.0)
    return df


@st.cache_data(ttl=3600, show_spinner=False)
def _load_principal_item_meta(company_id: str) -> pd.DataFrame:
    """Per-ItemID metadata for the principal: BottlesPerCase + MrpCaseRate +
    LastPurchase date. Drives the MRP-recode consolidation in Build mode —
    we need MrpCaseRate to pick the 'current' (highest-MRP) item in each
    consolidated group, and LastPurchase as a tie-breaker.

    Performance: LastPurchase is computed via a correlated subquery scoped
    to the last 24 months. The earlier all-history LEFT JOIN to TrVocItem
    + TrVocHead was pulling tens of millions of rows for principals with
    deep history and could time out on Streamlit Cloud."""
    type_ph = ",".join(str(t) for t in PURCHASE_TYPES)
    df = run_query(f"""
        SELECT im.ItemID,
               ISNULL(im.ItemDescription, im.ItemID)    AS ItemDescription,
               b.BrandName,
               im.BottlesPerCase,
               CAST(ISNULL(im.MrpCaseRate, 0) AS float) AS MrpCaseRate,
               (SELECT MAX(h.VoucherDate)
                  FROM TrVocItem vi
                  JOIN TrVocHead h
                    ON  h.TransTypeID  = vi.TransTypeID
                    AND h.VoucherNo    = vi.VoucherNo
                    AND h.FinancialYear = vi.FinancialYear
                 WHERE vi.ItemID = im.ItemID
                   AND h.TransTypeID IN ({type_ph})
                   AND h.Cancelled = 'N'
                   AND h.VoucherDate >= DATEADD(MONTH, -24, GETDATE())
               ) AS LastPurchase
        FROM MsItemMaster im
        JOIN MsBrandMaster b ON b.BrandID = im.BrandID
        WHERE b.CompanyID = ? AND im.ItemID LIKE 'I%'
    """, (company_id,))
    if not df.empty:
        df["BottlesPerCase"] = pd.to_numeric(df["BottlesPerCase"], errors="coerce") \
                                  .fillna(0).astype(int)
        df["MrpCaseRate"]   = pd.to_numeric(df["MrpCaseRate"], errors="coerce") \
                                  .fillna(0.0)
        df["LastPurchase"]  = pd.to_datetime(df["LastPurchase"], errors="coerce")
    return df


# Regexes for the SKU-base-key extraction (MRP-recode consolidation).
#   _MRP_DASH  strips "-NNN" / "-NNN X" (dash-separated MRP + optional 'N' tag)
#   _MRP_SPACE strips " NNN" (space-separated MRP, e.g. "QT (12) 5150")
#   _PAREN_RE  strips trailing "(K)", "(Hipster)", "(12)" variant markers
#
# The two MRP regexes are intentionally split: the dashed form allows an
# optional trailing letter ("-145 N" = NEW marker) which we want to strip;
# the spaced form must NOT consume a trailing letter because that would
# incorrectly eat pack sizes like "650M" or "750ML".
_MRP_DASH    = re.compile(r"\s*-\s*\d{2,5}(?:\s+[A-Z])?\s*$", re.IGNORECASE)
_MRP_SPACE   = re.compile(r"\s+\d{3,5}\s*$")
_PAREN_RE    = re.compile(r"\s*\([^\)]+\)\s*$")
# Trailing "NEW" / "-NEW" — added when the ERP re-codes a SKU. Strip so
# old + new MRP variants land in the same base key.
_NEW_SUFFIX  = re.compile(r"\s*[-\s]\s*NEW\s*$", re.IGNORECASE)


def _normalize_sku_key(desc: str) -> str:
    """Return a base key by stripping the trailing MRP token + parenthesised
    variant marker. Re-codes that differ ONLY in MRP collapse to the same
    key.

    Loops up to 3 times because some descriptions interleave the MRP and a
    paren, e.g. "CIROC VODKA QT (12)-4860" → strip "-4860" → "CIROC VODKA
    QT (12)" → strip "(12)" → "CIROC VODKA QT".

    Daman variants ("-D119" style) survive — the dash regex requires DIGITS
    immediately after the dash, so "-D119" is not stripped. Pack sizes
    like "650ML" / "750M" survive — the space-separated MRP regex only
    matches trailing digits with no following letter."""
    if not isinstance(desc, str):
        return ""
    s = re.sub(r"\s+", " ", desc.upper().strip())
    for _ in range(3):
        prev = s
        s = _MRP_DASH.sub("", s)
        s = _MRP_SPACE.sub("", s)
        s = _PAREN_RE.sub("", s)
        s = _NEW_SUFFIX.sub("", s)
        s = re.sub(r"\s*-\s*$", "", s).strip()
        if s == prev:
            break
    return re.sub(r"\s+", " ", s).strip()


def _normalize_brand_display(brand_name: str) -> str:
    """Clean display name for the catalog: strips '(12)' / '(Hipster)' /
    '-NEW' / ' NEW' so 'TANQUERAY LONDON GIN(12)' and 'TANQUERAY LONDON
    GIN-NEW' collapse to a single 'Tanqueray London Gin' card."""
    if not isinstance(brand_name, str):
        return ""
    s = brand_name.strip()
    s = _PAREN_RE.sub("", s)
    s = _NEW_SUFFIX.sub("", s)
    s = re.sub(r"\s+", " ", s).strip()
    # Title-case for nicer display (keep all-caps tokens like "JW" intact
    # by only title-casing words that have at least one lower-case letter
    # or are longer than 3 characters).
    def _cap(w):
        if len(w) <= 3 and w.isupper():
            return w
        return w[:1].upper() + w[1:].lower()
    return " ".join(_cap(w) for w in s.split())


def _consolidate_recodes(sku_df: pd.DataFrame,
                        item_meta: pd.DataFrame) -> pd.DataFrame:
    """Roll up MRP-recoded SKUs: items sharing the same base description +
    BottlesPerCase collapse into the CURRENT one (highest MrpCaseRate, then
    most-recent purchase, then highest ItemID). L3M / LYSM / Stock /
    Received MTD are SUMMED across the group and assigned to the current
    item; non-current items are dropped from the output."""
    if sku_df.empty:
        return sku_df

    meta = item_meta[["ItemID", "BottlesPerCase", "MrpCaseRate", "LastPurchase"]]
    df = sku_df.merge(meta, on="ItemID", how="left")
    df["BottlesPerCase"] = pd.to_numeric(df["BottlesPerCase"], errors="coerce") \
                              .fillna(0).astype(int)
    df["MrpCaseRate"]    = pd.to_numeric(df["MrpCaseRate"], errors="coerce") \
                              .fillna(0.0)
    df["LastPurchase"]   = pd.to_datetime(df["LastPurchase"], errors="coerce")

    df["_base"]  = df["ItemDescription"].apply(_normalize_sku_key)
    df["_group"] = df["_base"] + "||" + df["BottlesPerCase"].astype(str)
    # Pick the 'current' item per group: MOST RECENT PURCHASE first (the
    # variant we're actually buying NOW — handles MRP cuts correctly, e.g.
    # McDowell No1 NIP 220 → 210). Ties broken by highest MRP, then
    # highest ItemID (newest assigned code).
    df["_lp"] = df["LastPurchase"].fillna(pd.Timestamp("1900-01-01"))
    df = df.sort_values(
        ["_group", "_lp", "MrpCaseRate", "ItemID"],
        ascending=[True, False, False, False])
    df["_current"] = ~df["_group"].duplicated(keep="first")

    # Aggregate stats per group
    sum_cols = ["L3M_total", "LYSM", "Stock", "Received MTD", "L3M_active"]
    agg_cols = [c for c in sum_cols if c in df.columns]
    agg = df.groupby("_group")[agg_cols].sum()
    # Cap L3M_active at 3 (months in window) — variants may overlap.
    if "L3M_active" in agg.columns:
        agg["L3M_active"] = agg["L3M_active"].clip(upper=3)
    rollup_n = df.groupby("_group").size()

    # Replace stats on the current row; drop others
    out = df[df["_current"]].copy()
    for c in agg_cols:
        out[c] = out["_group"].map(agg[c])
    out["_rollup_n"] = out["_group"].map(rollup_n)
    if "L3M_active" in out.columns and "L3M_total" in out.columns:
        out["L3M_avg_adj"] = out.apply(
            lambda r: (r["L3M_total"] / r["L3M_active"])
                      if r["L3M_active"] > 0 else 0.0,
            axis=1)

    # Annotate the description so the user sees how many variants rolled up
    def _annotate(r):
        n = int(r.get("_rollup_n", 1))
        if n > 1:
            return f"{r['ItemDescription']}  (rolled up {n} MRP variants)"
        return r["ItemDescription"]
    out["ItemDescription"] = out.apply(_annotate, axis=1)

    drop_cols = ["_base", "_group", "_lp", "_current", "_rollup_n",
                 "BottlesPerCase", "MrpCaseRate", "LastPurchase"]
    return out.drop(columns=[c for c in drop_cols if c in out.columns])


@st.cache_data(ttl=1800, show_spinner=False)
def _load_principal_sku_history(company_id: str,
                                months_back: int = 13) -> pd.DataFrame:
    """Per (ItemID, ItemDescription, BrandName, yyyy-MM) cases for ONE
    principal — SKU-level granularity for the SKU-drill in Build mode."""
    type_ph = ",".join(str(t) for t in SALES_TYPES)
    df = run_query(f"""
        SELECT im.ItemID,
               ISNULL(im.ItemDescription, im.ItemID) AS ItemDescription,
               b.BrandName,
               FORMAT(h.VoucherDate, 'yyyy-MM') AS Mon,
               CAST(SUM(CAST(vi.TotalBottleQty AS decimal(18,4))
                        / NULLIF(im.BottlesPerCase, 0)) AS float) AS Cases
        FROM TrVocHead h
        JOIN TrVocItem vi
            ON  vi.TransTypeID = h.TransTypeID AND vi.VoucherNo = h.VoucherNo
            AND vi.ItemID LIKE 'I%'
            AND vi.FinancialYear = CASE
                WHEN MONTH(h.VoucherDate) >= 4
                THEN CAST(YEAR(h.VoucherDate) AS VARCHAR)+'-'+CAST(YEAR(h.VoucherDate)+1 AS VARCHAR)
                ELSE CAST(YEAR(h.VoucherDate)-1 AS VARCHAR)+'-'+CAST(YEAR(h.VoucherDate) AS VARCHAR)
              END
        JOIN MsItemMaster  im ON im.ItemID = vi.ItemID
        JOIN MsBrandMaster b  ON b.BrandID = im.BrandID
        WHERE h.TransTypeID IN ({type_ph}) AND h.Cancelled = 'N'
          AND b.CompanyID = ?
          AND h.VoucherDate >= DATEADD(MONTH, -{months_back}, GETDATE())
        GROUP BY im.ItemID, im.ItemDescription, b.BrandName,
                 FORMAT(h.VoucherDate, 'yyyy-MM')
    """, (company_id,))
    if not df.empty:
        df["Cases"] = pd.to_numeric(df["Cases"], errors="coerce").fillna(0.0)
    return df


def _sku_summary(cid: str, today: date,
                 stock_df: pd.DataFrame) -> pd.DataFrame:
    """Per-SKU L3M total / L3M avg/mo (stock-out adjusted) / LYSM / Stock,
    for a principal. One row per ItemID with sales OR stock in the principal."""
    hist = _load_principal_sku_history(cid, 13)

    # L3M / LYSM months (same logic as _brand_summary)
    prev_y, prev_m = today.year, today.month - 1
    if prev_m == 0:
        prev_m = 12; prev_y -= 1
    l3_months, y, m = [], prev_y, prev_m
    for _ in range(3):
        l3_months.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0: m = 12; y -= 1
    lysm_mon = f"{today.year-1:04d}-{today.month:02d}"

    # Per (ItemID, BrandName, ItemDescription) aggregates
    if not hist.empty:
        by = (hist
              .assign(IsL3M=hist["Mon"].isin(l3_months),
                      IsLYSM=hist["Mon"] == lysm_mon,
                      L3MAct=(hist["Mon"].isin(l3_months)
                              & (hist["Cases"] >= 1.0)).astype(int))
              .groupby(["ItemID", "ItemDescription", "BrandName"], as_index=False)
              .agg(L3M_total=("Cases", lambda s: s[hist.loc[s.index, "Mon"].isin(l3_months)].sum()),
                   LYSM=("Cases", lambda s: s[hist.loc[s.index, "Mon"] == lysm_mon].sum()),
                   L3M_active=("L3MAct", "sum")))
    else:
        by = pd.DataFrame(columns=["ItemID", "ItemDescription",
                                   "BrandName", "L3M_total", "LYSM",
                                   "L3M_active"])

    # Stock per ItemID (filter to this principal's brands)
    if not stock_df.empty and "ItemID" in stock_df.columns:
        # Find which BrandNames belong to this principal — using master
        master = _load_principal_brand_master(cid)
        principal_brands = set(master["BrandName"]) if not master.empty else set()
        s = stock_df[stock_df["BrandName"].isin(principal_brands)]
        stock_per_item = s.groupby(
            ["ItemID", "ItemDescription", "BrandName"]
        )["ClosingCases"].sum().reset_index().rename(
            columns={"ClosingCases": "Stock"})
        out = by.merge(stock_per_item,
                       on=["ItemID", "ItemDescription", "BrandName"],
                       how="outer")
    else:
        out = by.copy()
        out["Stock"] = 0.0

    for c in ("L3M_total", "LYSM", "L3M_active", "Stock"):
        if c in out.columns:
            out[c] = pd.to_numeric(out[c], errors="coerce").fillna(0.0)
        else:
            out[c] = 0.0
    out["L3M_avg_adj"] = out.apply(
        lambda r: (r["L3M_total"] / r["L3M_active"]) if r["L3M_active"] > 0 else 0.0,
        axis=1)
    return out


@st.cache_data(ttl=1800, show_spinner=False)
def _load_principal_brand_master(company_id: str) -> pd.DataFrame:
    df = run_query("""
        SELECT BrandID, BrandName FROM MsBrandMaster WHERE CompanyID = ?
    """, (company_id,))
    return df


def _brand_summary(cid: str, today: date, stock_df: pd.DataFrame) -> pd.DataFrame:
    """Per-brand summary for a principal: L3M total, L3M avg/mo (stock-out
    adjusted with <1 cs threshold), LYSM (full month a year ago), current
    stock. One row per BrandName in MsBrandMaster for the principal."""
    hist = _load_principal_brand_history(cid, 13)
    master = _load_principal_brand_master(cid)
    if master.empty:
        return pd.DataFrame()

    # Stock per brand (the principal's brands only)
    if not stock_df.empty:
        stk = (stock_df[stock_df["BrandName"].isin(master["BrandName"])]
               .groupby("BrandName")["ClosingCases"].sum().to_dict())
    else:
        stk = {}

    # L3M window = last 3 fully-elapsed months
    prev_y, prev_m = today.year, today.month - 1
    if prev_m == 0:
        prev_m = 12; prev_y -= 1
    l3_months = []
    y, m = prev_y, prev_m
    for _ in range(3):
        l3_months.append(f"{y:04d}-{m:02d}")
        m -= 1
        if m == 0: m = 12; y -= 1
    lysm_mon = f"{today.year-1:04d}-{today.month:02d}"

    rows = []
    by_brand_mon = hist.set_index(["BrandName", "Mon"])["Cases"].to_dict() \
                   if not hist.empty else {}
    for bn in master["BrandName"]:
        l3_vals = [float(by_brand_mon.get((bn, m), 0.0)) for m in l3_months]
        l3_active = sum(1 for v in l3_vals if v >= 1.0)
        l3_total = sum(l3_vals)
        l3_avg_n = l3_total / 3.0
        l3_avg_a = (l3_total / l3_active) if l3_active else 0.0
        lysm     = float(by_brand_mon.get((bn, lysm_mon), 0.0))
        rows.append({
            "BrandName":   bn,
            "Stock":       float(stk.get(bn, 0.0)),
            "L3M total":   l3_total,
            "L3M avg/mo (naive)": l3_avg_n,
            "L3M avg/mo (adj)":   l3_avg_a,
            "LYSM":        lysm,
            "L3M active mo": l3_active,
        })
    return pd.DataFrame(rows)


# ── Mode A: Analyze company's primary plan from an uploaded Excel ─────────
def _verdict(ask: float, stock: float, l3m_avg: float) -> str:
    if ask <= 0:
        return "Skip (no ask)"
    if stock >= ask + l3m_avg:
        return "Hold — over-stocked"
    if stock >= ask:
        return "Cover OK"
    gap = ask - stock
    if stock >= l3m_avg or l3m_avg <= 0:
        return f"Order ~{gap:,.0f} cs"
    return f"⚠️ Order ~{gap:,.0f} cs (low stock)"


def _match_brands(text: str, master_brands: list[str],
                  code_map: dict[str, list[str]]) -> list[str]:
    """Match the upload's brand-identifier text to master brand names.

    Priority: 1) lookup in code_map (for Diageo-style cryptic codes),
              2) case-insensitive substring of the text in master brand names.
    """
    t = (text or "").strip().upper()
    if not t:
        return []
    # Code-map first
    patterns = code_map.get(t) or code_map.get(t.replace(" ", ""))
    if patterns:
        return [b for b in master_brands
                if any(p in b.upper() for p in patterns)]
    # Fallback: substring match in either direction
    return [b for b in master_brands
            if t in b.upper() or b.upper() in t]


def _mode_analyze(stock_df: pd.DataFrame, cid: str, today: date) -> None:
    st.markdown("##### 📥 Analyze company's primary plan — upload Excel")
    st.caption(
        "Drop the company's primary-plan Excel. We'll match each brand to "
        "your master, compare the ask to current stock + L3M average + LYSM, "
        "and tell you what to order, hold, or skip."
    )
    up = st.file_uploader(
        "Primary-plan file (.xlsx)", type=["xlsx", "xls"],
        key=f"pp_upload_{cid}")
    if up is None:
        st.info("Upload a file to begin.")
        return

    try:
        # Read all rows ignoring headers — we'll detect headers below
        raw = pd.read_excel(up, header=None)
    except Exception as e:
        st.error(f"Could not read the file: {e}"); return

    # Auto-detect header row (the first row where >=2 cells are non-null
    # strings that look like header words)
    HDR_HINT = {"brand", "sku", "code", "cases", "qty", "quantity",
                "plan", "primary", "indent", "name"}
    hdr_row = None
    for i in range(min(10, len(raw))):
        cells = [str(c).strip().lower() for c in raw.iloc[i] if pd.notna(c)]
        if sum(any(h in c for h in HDR_HINT) for c in cells) >= 2:
            hdr_row = i; break
    if hdr_row is None:
        st.warning("Couldn't auto-detect a header row. Treating row 1 as headers.")
        hdr_row = 0
    df = raw.iloc[hdr_row + 1:].copy()
    df.columns = [str(c).strip() if pd.notna(c) else f"col{i}"
                  for i, c in enumerate(raw.iloc[hdr_row])]

    # Column mapping
    cols = list(df.columns)
    def _guess(keys):
        for c in cols:
            cl = c.lower()
            if any(k in cl for k in keys):
                return c
        return cols[0] if cols else None
    brand_default = _guess(["brand", "code", "name"])
    qty_default   = _guess(["plan", "primary", "case", "qty", "quantity", "indent"])

    c1, c2 = st.columns(2)
    with c1:
        brand_col = st.selectbox("Brand code/name column", cols,
                                 index=cols.index(brand_default) if brand_default in cols else 0,
                                 key=f"pp_bcol_{cid}")
    with c2:
        qty_col   = st.selectbox("Cases / quantity column", cols,
                                 index=cols.index(qty_default) if qty_default in cols else 0,
                                 key=f"pp_qcol_{cid}")

    df = df[[brand_col, qty_col]].copy()
    df.columns = ["BrandKey", "Cases"]
    df["BrandKey"] = df["BrandKey"].astype(str).str.strip()
    df["Cases"] = pd.to_numeric(df["Cases"], errors="coerce").fillna(0.0)
    df = df[df["BrandKey"].str.len() > 0]

    # Aggregate by brand key
    agg = df.groupby("BrandKey", as_index=False)["Cases"].sum() \
            .rename(columns={"Cases": "Ask"})
    if agg.empty:
        st.info("File parsed but no usable rows found."); return

    # Match each key to master brands
    summary = _brand_summary(cid, today, stock_df)
    if summary.empty:
        st.error("Couldn't load brand master for this principal."); return
    master_brands = summary["BrandName"].tolist()

    code_map = DIAGEO_CODE_MAP if cid == "C00040" else {}

    rows = []
    for key, ask in zip(agg["BrandKey"], agg["Ask"]):
        matched = _match_brands(key, master_brands, code_map)
        if matched:
            sub = summary[summary["BrandName"].isin(matched)]
            stock = float(sub["Stock"].sum())
            l3m   = float(sub["L3M total"].sum())
            l3_avg = float(sub["L3M avg/mo (adj)"].sum())
            lysm  = float(sub["LYSM"].sum())
            verdict = _verdict(ask, stock, l3_avg)
            mapped = "; ".join(matched[:3]) + (
                f" (+{len(matched)-3} more)" if len(matched) > 3 else "")
        else:
            stock = l3m = l3_avg = lysm = 0.0
            verdict = "⚠️ Unmapped — check"
            mapped = "(none)"
        rows.append({
            "Brand key": key, "Mapped brand(s)": mapped, "Ask (cs)": ask,
            "Stock": stock, "L3M total": l3m, "L3M avg/mo": l3_avg,
            "LYSM": lysm, "Verdict": verdict,
        })
    out = pd.DataFrame(rows).sort_values("Ask (cs)", ascending=False)

    # Headline
    total_ask = float(out["Ask (cs)"].sum())
    n_order = int(out["Verdict"].str.startswith(("Order", "⚠️ Order")).sum())
    n_hold  = int((out["Verdict"] == "Hold — over-stocked").sum())
    n_cover = int((out["Verdict"] == "Cover OK").sum())
    n_un    = int((out["Verdict"] == "⚠️ Unmapped — check").sum())
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total ask", f"{total_ask:,.0f} cs")
    c2.metric("Orders", f"{n_order} brands")
    c3.metric("Hold", f"{n_hold} brands")
    c4.metric("Unmapped", f"{n_un} brands")

    def _flag(v):
        s = str(v)
        if "⚠️" in s:        return "color:#dc2626;font-weight:700"
        if "Hold" in s:      return "color:#16a34a;font-weight:600"
        if "Cover OK" in s:  return "color:#16a34a;font-weight:600"
        if "Order" in s:     return "color:#b45309;font-weight:600"
        return ""

    st.dataframe(
        out.style.format({c: "{:,.1f}" for c in
                         ["Ask (cs)", "Stock", "L3M total", "L3M avg/mo", "LYSM"]})
           .map(_flag, subset=["Verdict"]),
        use_container_width=True, hide_index=True, height=480)
    st.download_button(
        "⬇️ Download annotated plan",
        out.to_csv(index=False).encode("utf-8-sig"),
        file_name=f"primary_plan_analysis_{cid}_{today.isoformat()}.csv",
        mime="text/csv", key=f"pp_dl_{cid}")


# ── Mode B: Build OUR plan from a total target ─────────────────────────────
def _mode_build(stock_df: pd.DataFrame, cid: str, today: date,
                principal_name: str) -> None:
    from src.targets import load_primary_plan, save_primary_plan
    st.markdown("##### 📝 Build our primary plan — total target → per-brand split")
    st.caption(
        "Enter the total cases you're planning to order. We allocate it "
        "across brands using your real sales history. The **planned indent "
        "is over and above current stock** — Stock is shown for context "
        "only and is NOT subtracted from the indent. **Indent = Allocated − "
        "Received MTD**, so it reflects the balance of the planned order "
        "that still has to come from the company."
    )

    # Persist the total per (principal, month) so it survives reloads.
    month_str = f"{today.year:04d}-{today.month:02d}"
    saved = load_primary_plan(cid, month_str)
    default_total = float(saved["total"]) if saved else 0.0

    c1, c2, c3 = st.columns([1, 1, 1])
    with c1:
        total = st.number_input(
            "Total target (cases)", min_value=0.0, value=default_total,
            step=100.0, key=f"pp_build_total_{cid}_{month_str}")
        # Save when the value changes
        if total > 0 and abs(total - default_total) > 0.001:
            save_primary_plan(cid, month_str, total)
            st.caption(f"💾 Saved {total:,.0f} cs for {principal_name} "
                       f"({month_str}). Persists across reloads until you "
                       f"change it.")
        elif saved:
            st.caption(f"💾 Loaded saved target of {default_total:,.0f} cs "
                       f"(set on {saved.get('updated_at','?')}).")
    with c2:
        method = st.radio(
            "Distribution basis",
            ["L3M share", "LYSM share", "Blend (50/50)"],
            horizontal=False, key=f"pp_build_method_{cid}")
    with c3:
        granularity = st.radio(
            "Granularity",
            ["Brand-level", "SKU-level"],
            horizontal=False, key=f"pp_build_gran_{cid}",
            help="Brand-level allocates the total across brands. "
                 "SKU-level then breaks each brand's share across its SKUs "
                 "by their own L3M share inside the brand.")

    if total <= 0:
        st.info("Enter a total target to see the suggested per-brand split.")
        return

    summary = _brand_summary(cid, today, stock_df)
    if summary.empty:
        st.error("Couldn't load brand data."); return

    # Weights
    w_l3m  = summary["L3M total"].clip(lower=0)
    w_lysm = summary["LYSM"].clip(lower=0)
    if method == "L3M share":
        w = w_l3m
    elif method == "LYSM share":
        w = w_lysm
    else:
        # 50/50 blend on share, then apply to total
        s_l3m  = w_l3m.sum()
        s_lysm = w_lysm.sum()
        share_l3m  = (w_l3m  / s_l3m)  if s_l3m  > 0 else pd.Series(0, index=summary.index)
        share_lysm = (w_lysm / s_lysm) if s_lysm > 0 else pd.Series(0, index=summary.index)
        share = (share_l3m + share_lysm) / 2.0
        # Renormalise (handles brands with only one source)
        if share.sum() > 0:
            share = share / share.sum()
        summary["Share %"] = (share * 100.0).round(2)
        summary["Allocated"] = (share * total).round(1)
        w = None

    if w is not None:
        s = w.sum()
        share = (w / s) if s > 0 else pd.Series(0, index=summary.index)
        summary["Share %"]   = (share * 100.0).round(2)
        summary["Allocated"] = (share * total).round(1)

    # Received MTD (already arrived this month) — per brand
    rec_mtd = _load_principal_received_mtd(cid, today)
    if not rec_mtd.empty:
        rec_brand = rec_mtd.groupby("BrandName")["Cases"].sum().to_dict()
    else:
        rec_brand = {}
    summary["Received MTD"] = summary["BrandName"].map(rec_brand).fillna(0.0)

    # Indent = balance of the planned order still pending from the company.
    # The plan is OVER AND ABOVE current stock, so Stock is NOT subtracted.
    # Received MTD is subtracted because those cases are already part of
    # this month's deliveries against the plan.
    summary["Indent"] = (summary["Allocated"] - summary["Received MTD"]) \
                        .clip(lower=0).round(1)

    # Show only brands with any historic activity OR an allocation
    keep = (summary["L3M total"] > 0) | (summary["LYSM"] > 0) | (summary["Stock"] > 0) | (summary["Allocated"] > 0) | (summary["Received MTD"] > 0)
    out = summary[keep].sort_values("Allocated", ascending=False).copy()

    # Headline
    c1, c2, c3 = st.columns(3)
    c1.metric("Allocated total", f"{out['Allocated'].sum():,.0f} cs")
    c2.metric("Indent total (after stock)", f"{out['Indent'].sum():,.0f} cs")
    c3.metric("Stock offset", f"{(out['Allocated'].sum() - out['Indent'].sum()):,.0f} cs")

    disp = out[["BrandName", "L3M avg/mo (adj)", "LYSM", "Share %",
                "Allocated", "Stock", "Received MTD", "Indent"]].rename(columns={
        "BrandName": "Brand", "L3M avg/mo (adj)": "L3M/mo (adj)"})

    def _style_indent(v):
        try:
            return ("color:#dc2626;font-weight:700" if v > 0
                    else "color:#16a34a")
        except Exception:
            return ""

    st.dataframe(
        disp.style.format({c: "{:,.1f}" for c in
                          ["L3M/mo (adj)", "LYSM", "Allocated", "Stock",
                           "Received MTD", "Indent"]}
                          | {"Share %": "{:.2f}%"})
            .map(_style_indent, subset=["Indent"]),
        use_container_width=True, hide_index=True, height=520)
    st.download_button(
        "⬇️ Download brand split",
        disp.to_csv(index=False).encode("utf-8-sig"),
        file_name=f"primary_plan_build_{cid}_{today.isoformat()}.csv",
        mime="text/csv", key=f"pp_build_dl_{cid}")

    # ── SKU-level drill (optional) ─────────────────────────────────────
    if granularity != "SKU-level":
        return
    st.markdown("---")
    st.markdown("##### SKU-level breakdown")
    st.caption(
        "Each brand's allocation, split across its SKUs by the SKU's L3M "
        "share inside the brand. Brands with no L3M history use a flat "
        "split across their SKUs. Stock and Received MTD are matched "
        "per-ItemID. **Indent = Allocated − Received MTD** — over and "
        "above current stock."
    )

    consolidate = st.checkbox(
        "Consolidate MRP re-codes (KF Strong 195+200 → 200 with combined history)",
        value=True, key=f"pp_consolidate_{cid}",
        help="When MRP changes, the ERP issues a new ItemID. The old "
             "ItemID still has L3M / LYSM sales but no more stock will "
             "arrive of it.\n"
             "ON  → roll up all MRP variants of the same physical SKU into "
             "the current item (highest MRP, latest purchase). Description "
             "shows '(rolled up N MRP variants)'.\n"
             "OFF → show every ItemID separately (raw view).")

    try:
        sku = _sku_summary(cid, today, stock_df)
        if sku.empty:
            st.info("No SKU history available for this principal.")
            return

        # Per-SKU Received MTD — load before consolidation so the rollup includes it
        rec_sku_df = _load_principal_received_mtd(cid, today)
        rec_by_item = (dict(zip(rec_sku_df["ItemID"], rec_sku_df["Cases"]))
                       if not rec_sku_df.empty else {})
        sku["Received MTD"] = sku["ItemID"].map(rec_by_item).fillna(0.0)

        # MRP-recode consolidation (BEFORE allocation, so shares use the
        # combined L3M of all variants)
        if consolidate:
            meta = _load_principal_item_meta(cid)
            before_n = len(sku)
            sku = _consolidate_recodes(sku, meta)
            after_n = len(sku)
            if before_n != after_n:
                st.caption(
                    f"🔁 Consolidated {before_n - after_n} MRP re-coded SKUs "
                    f"into their current variants ({before_n} → {after_n} rows).")
    except Exception as exc:
        st.error(
            "⚠️ **Could not load SKU breakdown** — the SQL Server likely "
            "dropped the connection. Click 🔄 Refresh at the top-right (or "
            "Ctrl+Shift+R) to retry.\n\n"
            f"_Technical: `{type(exc).__name__}: {str(exc)[:200]}`_")
        return

    # Map brand → allocated cases
    brand_alloc = dict(zip(out["BrandName"], out["Allocated"]))
    sku = sku[sku["BrandName"].isin(brand_alloc.keys())].copy()
    if sku.empty:
        st.info("No SKUs to allocate (no brand received a positive allocation).")
        return

    # SKU share within brand → SKU allocation (using post-consolidation L3M)
    sku["BrandAlloc"]  = sku["BrandName"].map(brand_alloc).fillna(0.0)
    sku["BrandL3M"]    = sku.groupby("BrandName")["L3M_total"].transform("sum")
    sku["SkuPerBrand"] = sku.groupby("BrandName")["ItemID"].transform("count")
    def _sku_share(row):
        if row["BrandL3M"] > 0:
            return row["L3M_total"] / row["BrandL3M"]
        if row["SkuPerBrand"] > 0:
            return 1.0 / row["SkuPerBrand"]    # flat fallback
        return 0.0
    sku["Share"]    = sku.apply(_sku_share, axis=1)
    sku["Allocated"] = (sku["BrandAlloc"] * sku["Share"]).round(1)

    # Indent = balance still to come from the company (planned − already
    # received this month). Stock is NOT subtracted — the planned indent is
    # over and above current stock.
    sku["Indent"] = (sku["Allocated"] - sku["Received MTD"]).clip(lower=0).round(1)

    # Only show SKUs whose brand got an allocation > 0 OR which carry stock
    keep_sku = (sku["BrandAlloc"] > 0) | (sku["Stock"] > 0)
    sku = sku[keep_sku].copy()
    if sku.empty:
        st.info("No SKUs to display."); return

    sku_disp = (sku[["BrandName", "ItemDescription", "L3M_avg_adj", "LYSM",
                     "Share", "BrandAlloc", "Allocated", "Stock",
                     "Received MTD", "Indent"]]
                .sort_values(["BrandName", "Allocated"], ascending=[True, False])
                .rename(columns={"BrandName": "Brand", "ItemDescription": "SKU",
                                 "L3M_avg_adj": "L3M/mo", "Share": "Share in brand",
                                 "BrandAlloc": "Brand alloc"}))

    # Headline
    c1, c2, c3 = st.columns(3)
    c1.metric("SKUs in plan", f"{len(sku_disp):,}")
    c2.metric("SKU indent total", f"{sku_disp['Indent'].sum():,.0f} cs")
    c3.metric("Indent rows >0", f"{(sku_disp['Indent']>0).sum():,}")

    st.dataframe(
        sku_disp.style.format({
            "L3M/mo": "{:,.2f}", "LYSM": "{:,.1f}",
            "Brand alloc": "{:,.1f}", "Allocated": "{:,.1f}",
            "Stock": "{:,.1f}", "Received MTD": "{:,.1f}",
            "Indent": "{:,.1f}",
            "Share in brand": "{:.1%}",
        }).map(_style_indent, subset=["Indent"]),
        use_container_width=True, hide_index=True, height=620)
    st.download_button(
        "⬇️ Download SKU split",
        sku_disp.to_csv(index=False).encode("utf-8-sig"),
        file_name=f"primary_plan_build_sku_{cid}_{today.isoformat()}.csv",
        mime="text/csv", key=f"pp_build_sku_dl_{cid}")


# ─────────────────────────────────────────────────────────────────────────────
# Brand Catalog — auto-updated rate sheet for sharing externally
# ─────────────────────────────────────────────────────────────────────────────

# Sales rate (outlet's purchase rate) is stored authoritatively on
# MsItemRates.SaleCaseRate / SaleBottleRate. The table is dated (ApplyDate)
# and can carry multiple historical entries per ItemID, so we always pick
# the LATEST row (max id_key) per item.

@st.cache_data(ttl=3600, show_spinner=False)
def _load_brand_catalog_master() -> pd.DataFrame:
    """Per-ItemID master data driving the Catalog: BottlesPerCase, MRP per
    case + bottle, AUTHORITATIVE Sales Rate (from MsItemRates — the latest
    entry per ItemID), LiquorSize, and LastPurchase (date of the most
    recent purchase voucher, last 24 months). LastPurchase drives the
    MRP-recode consolidation so we pick the variant we're actually buying
    NOW — handles MRP cuts (McDowell No1 NIP 220 → 210) as well as MRP
    increases.

    Valuation/landed rates are deliberately NOT pulled — they're our
    internal cost, not for the customer sheet."""
    type_ph = ",".join(str(t) for t in PURCHASE_TYPES)
    df = run_query(f"""
        SELECT
            im.ItemID,
            ISNULL(im.ItemDescription, im.ItemID)        AS ItemDescription,
            im.BottlesPerCase,
            CAST(ISNULL(im.MrpCaseRate, 0)         AS float) AS MrpCase,
            CAST(ISNULL(im.MrpBottRate, 0)         AS float) AS MrpBottle,
            CAST(ISNULL(r.SaleCaseRate, 0)         AS float) AS SaleCase,
            CAST(ISNULL(r.SaleBottleRate, 0)       AS float) AS SaleBottle,
            ISNULL(im.LiquorSize, '')                    AS LiquorSize,
            b.BrandName,
            b.CompanyID,
            (SELECT MAX(h.VoucherDate)
                FROM TrVocItem vi
                JOIN TrVocHead h
                    ON  h.TransTypeID  = vi.TransTypeID
                    AND h.VoucherNo    = vi.VoucherNo
                    AND h.FinancialYear = vi.FinancialYear
                WHERE vi.ItemID = im.ItemID
                  AND h.TransTypeID IN ({type_ph})
                  AND h.Cancelled = 'N'
                  AND h.VoucherDate >= DATEADD(MONTH, -24, GETDATE())
            ) AS LastPurchase
        FROM MsItemMaster im
        JOIN MsBrandMaster b ON b.BrandID = im.BrandID
        LEFT JOIN (
            SELECT ItemID, SaleCaseRate, SaleBottleRate FROM (
                SELECT ItemID, SaleCaseRate, SaleBottleRate,
                       ROW_NUMBER() OVER
                           (PARTITION BY ItemID
                            ORDER BY ApplyDate DESC, id_key DESC) AS rn
                FROM MsItemRates
            ) x WHERE rn = 1
        ) r ON r.ItemID = im.ItemID
        WHERE im.ItemID LIKE 'I%'
          AND b.CompanyID IN ('C00025','C00039','C00040','C00056')
    """)
    if not df.empty:
        df["BottlesPerCase"] = pd.to_numeric(df["BottlesPerCase"], errors="coerce") \
                                   .fillna(0).astype(int)
        df["LastPurchase"]   = pd.to_datetime(df["LastPurchase"], errors="coerce")
    return df


@st.cache_data(ttl=900, show_spinner=False)
def _load_recently_active_items(months_back: int = 6) -> set[str]:
    """ItemIDs that appeared on any voucher (sales or purchase) in the last
    N months. Combined with current stock, defines 'active' for the catalog."""
    df = run_query(f"""
        SELECT DISTINCT vi.ItemID
        FROM TrVocItem vi
        JOIN TrVocHead h
            ON  h.TransTypeID = vi.TransTypeID
            AND h.VoucherNo   = vi.VoucherNo
            AND h.FinancialYear = vi.FinancialYear
        WHERE h.Cancelled = 'N' AND vi.ItemID LIKE 'I%'
          AND h.VoucherDate >= DATEADD(MONTH, -{months_back}, GETDATE())
    """)
    return set(df["ItemID"]) if not df.empty else set()


def _consolidate_catalog_recodes(cat: pd.DataFrame) -> pd.DataFrame:
    """Keep only the CURRENT MRP variant per physical SKU. Groups by
    (normalised base description, BottlesPerCase). 'Current' = the variant
    we're actually buying NOW, identified as the most-recent LastPurchase;
    ties broken by highest MrpCase, then highest ItemID.

    Crucially this handles MRP CUTS (McDowell No1 NIP 220 → 210) correctly
    — the 210 wins because it's the one being purchased, even though 220
    is the higher MRP. Where no purchase data exists (e.g. brand-new SKU
    with no in-window purchases yet), MrpCase becomes the primary key."""
    if cat.empty:
        return cat
    df = cat.copy()
    df["_base"]  = df["ItemDescription"].apply(_normalize_sku_key)
    df["_group"] = df["_base"] + "||" + df["BottlesPerCase"].astype(str)
    # Sentinel so NaN LastPurchase sorts to the bottom, not the top
    df["_lp"] = df["LastPurchase"].fillna(pd.Timestamp("1900-01-01"))
    df = df.sort_values(
        ["_group", "_lp", "MrpCase", "ItemID"],
        ascending=[True, False, False, False])
    df = df[~df["_group"].duplicated(keep="first")]
    return df.drop(columns=["_base", "_group", "_lp"])


# ── Category classifier for the catalog ─────────────────────────────────
# Owner-defined order: numeric prefix sorts the bands in the right sequence,
# and the prefix is stripped before display so the user sees clean names.
def _catalog_category(cid: str, brand_name: str, mrp_bottle: float) -> str:
    b = (brand_name or "").upper()
    if cid == "C00039":     # United Breweries
        if "CANNON" in b or "LONDON PIL" in b:
            return "1. Economy"
        if "ULTRA MAX" in b:
            return "5. Ultra Max"
        if "ULTRA" in b:
            return "4. Ultra"
        if "HEINEKEN" in b or "AMSTEL" in b:
            return "3. Premium (Heineken / Amstel)"
        if "KING FISHER" in b or "KINGFISHER" in b:
            return "2. Mainstream"
        return "9. Other"
    if cid == "C00025":     # United Spirits
        if "SIGNATURE" in b or "ANTIQUITY" in b:
            return "1. Upper Prestige"
        if "MCDOWELL" in b:
            return "2. Mid Prestige (McDowell's No 1)"
        return "9. Other"
    if cid == "C00040":     # Diageo
        if "SMIRNOFF" in b:
            return "1. Smirnoff (India-made / BII)"
        # Everything else is BIO (Bottled-In-Origin / imported)
        return ("2. BIO Premium (MRP > ₹3,500)"
                if (mrp_bottle or 0) > 3500
                else "3. BIO Standard (MRP ≤ ₹3,500)")
    if cid == "C00056":     # Brown-Forman
        return "1. Premium"
    return "9. Other"


def _strip_category_prefix(cat: str) -> str:
    """'1. Economy' → 'Economy' — drops the sort-prefix for display."""
    return re.sub(r"^\d+\.\s*", "", cat or "")


def _format_pack(row: pd.Series) -> str:
    """'12 × 650 ML' style pack label."""
    size = str(row.get("LiquorSize", "") or "").strip()
    bpc  = int(row.get("BottlesPerCase", 0) or 0)
    if bpc <= 0 and not size:
        return ""
    if bpc <= 0:
        return size
    if not size:
        return f"{bpc} × ?"
    # Normalise '650 ML' / '650ML' / '650 ml' → '650 ML'
    s = size.upper().replace("MILILITER", "ML").replace("MILLILITER", "ML")
    s = re.sub(r"\s+", " ", s).strip()
    return f"{bpc} × {s}"


def _build_catalog_xlsx(cat: pd.DataFrame) -> bytes:
    """Print-ready landscape Excel: one sheet per principal, brand-grouped."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO
    from datetime import date as _date

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    thin = Side(border_style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill   = PatternFill(start_color="1B4F72", end_color="1B4F72",
                                fill_type="solid")
    category_fill = PatternFill(start_color="1B4F72", end_color="1B4F72",
                                fill_type="solid")   # navy band for categories
    brand_fill    = PatternFill(start_color="FAE6BA", end_color="FAE6BA",
                                fill_type="solid")   # soft amber for brands

    today_str = _date.today().strftime("%d %b %Y")

    for principal, prin_df in cat.groupby("Principal"):
        ws = wb.create_sheet(principal[:31])
        # Top title bar
        ws.merge_cells("A1:E1")
        ws["A1"] = "Kranti Wines Pvt. Ltd."
        ws["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
        ws["A1"].fill = header_fill
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        ws.merge_cells("A2:E2")
        ws["A2"] = f"{principal} — Current MRP & Landing Rates (as of {today_str})"
        ws["A2"].font = Font(name="Arial", size=12, italic=True, color="1B4F72")
        ws["A2"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 22

        # Column headers — no separate SKU column; Brand + Pack identifies.
        hdr = ["Brand", "Pack", "Bottles / Case",
               "MRP / Case (₹)", "MRP / Bottle (₹)",
               "Sales Rate / Case (₹)", "Sales Rate / Bottle (₹)"]
        ncols = len(hdr)
        # Re-merge the title row to span all columns
        end_col = get_column_letter(ncols)
        ws.unmerge_cells("A1:E1"); ws.merge_cells(f"A1:{end_col}1")
        ws.unmerge_cells("A2:E2"); ws.merge_cells(f"A2:{end_col}2")
        for ci, h in enumerate(hdr, 1):
            c = ws.cell(row=4, column=ci, value=h)
            c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center",
                                    wrap_text=True)
            c.border = border
        ws.row_dimensions[4].height = 32

        r = 5
        for category in sorted(prin_df["Category"].unique()):
            cat_df = prin_df[prin_df["Category"] == category]
            # Category band — full-width navy header
            ws.merge_cells(start_row=r, start_column=1,
                           end_row=r, end_column=ncols)
            cc = ws.cell(row=r, column=1,
                         value=_strip_category_prefix(category))
            cc.font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
            cc.fill = category_fill
            cc.alignment = Alignment(horizontal="left", vertical="center",
                                     indent=1)
            cc.border = border
            ws.row_dimensions[r].height = 22
            r += 1
            for brand, brand_df in cat_df.groupby("BrandDisplay"):
                # Vertically merge the Brand cell across all of this brand's
                # pack rows so the brand name shows ONCE per group, packs
                # listed underneath — same shape as the manual KW Rates sheet.
                pack_rows = brand_df.sort_values("MrpCase", ascending=False)
                first_r = r
                for _, row in pack_rows.iterrows():
                    vals = [
                        brand,    # Brand name (only the first row's cell is visible
                                  # after the vertical merge below)
                        _format_pack(row),
                        int(row.get("BottlesPerCase", 0) or 0),
                        float(row.get("MrpCase",    0) or 0),
                        float(row.get("MrpBottle",  0) or 0),
                        float(row.get("SaleCase",   0) or 0),
                        float(row.get("SaleBottle", 0) or 0),
                    ]
                    for ci, v in enumerate(vals, 1):
                        c = ws.cell(row=r, column=ci, value=v)
                        c.font = Font(name="Arial", size=10)
                        c.border = border
                        if ci == 1:
                            c.fill = brand_fill
                            c.font = Font(name="Arial", size=10, bold=True)
                            c.alignment = Alignment(horizontal="left",
                                                    vertical="center",
                                                    indent=1, wrap_text=True)
                        elif ci == 3:
                            c.number_format = "#,##0"
                            c.alignment = Alignment(horizontal="center")
                        elif ci >= 4:
                            c.number_format = "#,##0"
                            c.alignment = Alignment(horizontal="right")
                    r += 1
                # Merge Brand cell vertically across this brand's packs
                if r - first_r > 1:
                    ws.merge_cells(start_row=first_r, end_row=r - 1,
                                   start_column=1, end_column=1)
                    # Re-apply alignment to the merged top-left cell
                    bc = ws.cell(row=first_r, column=1)
                    bc.alignment = Alignment(horizontal="left",
                                             vertical="center",
                                             indent=1, wrap_text=True)

        # Column widths — no SKU column anymore
        widths = [32, 18, 11, 16, 16, 18, 18]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = "A5"
        # Print setup — landscape, fit-to-page
        ws.page_setup.orientation = "landscape"
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.print_options.horizontalCentered = True
        ws.page_margins.left = 0.3
        ws.page_margins.right = 0.3
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _section_brand_catalog(stock_df: pd.DataFrame) -> None:
    st.subheader("📋 Brand Catalog — auto-updated rate sheet")
    st.caption(
        "Every active SKU per principal/brand with current MRP and landing "
        "rate. **Updates automatically** when MRPs change in the ERP — no "
        "more manual maintenance. Download the Excel and Print → Save as PDF "
        "(landscape, fit-to-page) to share with customers asking which "
        "brands you distribute."
    )

    # Filter row
    c1, c2 = st.columns([1, 2])
    with c1:
        principal_filter = st.selectbox(
            "Principal",
            ["All", "United Breweries", "United Spirits", "Diageo", "Brown-Forman"],
            key="cat_principal_filter")
    with c2:
        st.caption(
            "Active = had purchases/sales in the last 6 months OR carries "
            "current stock. MRP variants are consolidated to the highest "
            "(current) one, so re-codes like KFS 195 + 200 show once.")

    try:
        cat = _load_brand_catalog_master()
        active = _load_recently_active_items(6)
    except Exception as exc:
        st.error(
            "⚠️ Could not load the catalog — the SQL Server may have dropped "
            "the connection. Click 🔄 Refresh at the top-right and try again.\n\n"
            f"_Technical: `{type(exc).__name__}: {str(exc)[:200]}`_")
        return

    if cat.empty:
        st.info("No catalog data returned."); return

    # Active = (had recent voucher activity) OR (carries stock now)
    in_stock = (set(stock_df["ItemID"]) if not stock_df.empty
                and "ItemID" in stock_df.columns else set())
    cat = cat[cat["ItemID"].isin(active | in_stock)].copy()
    if cat.empty:
        st.info("No active SKUs in the window."); return

    # Consolidate MRP re-codes — keep only the highest-MRP variant per SKU
    before = len(cat)
    cat = _consolidate_catalog_recodes(cat)
    after = len(cat)

    # Principal name + filter
    cat["Principal"] = cat["CompanyID"].map(_PRINCIPAL_NAMES).fillna("Other")
    # Brand display: collapse "Tanqueray London Gin(12)" + "...-NEW" → one
    cat["BrandDisplay"] = cat["BrandName"].apply(_normalize_brand_display)
    # Category classifier (owner-defined groupings — see _catalog_category)
    cat["Category"] = cat.apply(
        lambda r: _catalog_category(r["CompanyID"], r["BrandName"],
                                    r.get("MrpBottle", 0) or 0),
        axis=1)
    if principal_filter != "All":
        cat = cat[cat["Principal"] == principal_filter]
    if cat.empty:
        st.info(f"No active SKUs for {principal_filter}."); return

    # KPIs
    k1, k2, k3 = st.columns(3)
    k1.metric("Principals", cat["Principal"].nunique())
    k2.metric("Brands",     cat["BrandName"].nunique())
    k3.metric("SKUs",       len(cat),
              help=f"Consolidated from {before} raw items "
                   f"(de-duplicated {before-after} MRP re-codes).")

    # Download
    try:
        xlsx_bytes = _build_catalog_xlsx(cat)
        st.download_button(
            "⬇️ Download rate sheet (Excel — print-ready, landscape)",
            xlsx_bytes,
            file_name=f"KW_Rates_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="cat_xlsx_dl")
    except Exception as exc:
        st.warning(f"Excel generation failed: {exc}")

    st.divider()

    # On-screen display: per principal → per category → per brand (display)
    for prin in sorted(cat["Principal"].unique()):
        prin_df = cat[cat["Principal"] == prin]
        n_brands = prin_df["BrandDisplay"].nunique()
        st.markdown(f"### {prin}  ·  {len(prin_df)} SKUs · {n_brands} brands")
        for category in sorted(prin_df["Category"].unique()):
            cat_df = prin_df[prin_df["Category"] == category]
            st.markdown(f"##### {_strip_category_prefix(category)}  "
                        f"<span style='color:#888;font-weight:normal'>"
                        f"· {len(cat_df)} SKUs</span>",
                        unsafe_allow_html=True)
            for brand in sorted(cat_df["BrandDisplay"].unique()):
                brand_df = cat_df[cat_df["BrandDisplay"] == brand]
                with st.expander(f"**{brand}**  ·  {len(brand_df)} pack"
                                 f"{'s' if len(brand_df) != 1 else ''}",
                                 expanded=False):
                    disp = brand_df.copy()
                    disp["Pack"] = disp.apply(_format_pack, axis=1)
                    show = (disp[["Pack", "BottlesPerCase",
                                  "MrpCase", "MrpBottle",
                                  "SaleCase", "SaleBottle"]]
                            .rename(columns={
                                "BottlesPerCase":   "Bottles / Case",
                                "MrpCase":          "MRP / Case (₹)",
                                "MrpBottle":        "MRP / Bottle (₹)",
                                "SaleCase":         "Sales Rate / Case (₹)",
                                "SaleBottle":       "Sales Rate / Bottle (₹)"})
                            .sort_values("MRP / Case (₹)", ascending=False))
                    st.dataframe(
                        show.style.format({
                            "Bottles / Case":           "{:,.0f}",
                            "MRP / Case (₹)":           "₹{:,.0f}",
                            "MRP / Bottle (₹)":         "₹{:,.0f}",
                            "Sales Rate / Case (₹)":    "₹{:,.0f}",
                            "Sales Rate / Bottle (₹)":  "₹{:,.0f}",
                        }),
                        use_container_width=True, hide_index=True)


def _section_primary_plan(stock_df: pd.DataFrame, as_of: date) -> None:
    st.subheader("📦 Primary Plan / Indent")
    st.caption(
        "One place for every primary-order workflow: analyse a plan the "
        "company sent you, build a plan to send back, or phase Brindco's "
        "annual transition across 12 months."
    )

    mode = st.radio(
        "Mode",
        ["📥 Analyze company plan (upload Excel)",
         "📝 Build our plan (manual target)",
         "🎯 Brindco annual transition"],
        horizontal=False, key="pp_mode")

    if mode.startswith("🎯"):
        _section_target_phasing(stock_df, as_of)
        return

    principal_name = st.selectbox(
        "Principal", list(_PRINCIPAL_CID_MAP.keys()),
        index=1,    # United Breweries default
        key="pp_principal")
    cid = _PRINCIPAL_CID_MAP[principal_name]

    if mode.startswith("📥"):
        _mode_analyze(stock_df, cid, as_of)
    else:
        _mode_build(stock_df, cid, as_of, principal_name)


# ─────────────────────────────────────────────────────────────────────────────
# Annual-target phasing — feeds an annual case target through a brand family's
# historical seasonality (stock-out-adjusted) to suggest a monthly indent.
# Designed for the Brindco transition (Jun-2026) but generic to any preset.
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data(ttl=1800, show_spinner=False)
def _load_family_monthly(months_back: int = 13) -> pd.DataFrame:
    """Per (BrandName, yyyy-MM) cases sold over the last N months.
    Family roll-up happens in pandas (one query covers all families).
    Spirits don't have kegs so the plain bottles/case math is fine."""
    type_ph = ",".join(str(t) for t in SALES_TYPES)
    df = run_query(f"""
        SELECT b.BrandName,
               FORMAT(h.VoucherDate, 'yyyy-MM') AS Mon,
               CAST(SUM(CAST(vi.TotalBottleQty AS decimal(18,4))
                        / NULLIF(im.BottlesPerCase, 0)) AS float) AS Cases
        FROM TrVocHead h
        JOIN TrVocItem vi
            ON  vi.TransTypeID = h.TransTypeID AND vi.VoucherNo = h.VoucherNo
            AND vi.ItemID LIKE 'I%'
            AND vi.FinancialYear = CASE
                WHEN MONTH(h.VoucherDate) >= 4
                THEN CAST(YEAR(h.VoucherDate) AS VARCHAR)+'-'+CAST(YEAR(h.VoucherDate)+1 AS VARCHAR)
                ELSE CAST(YEAR(h.VoucherDate)-1 AS VARCHAR)+'-'+CAST(YEAR(h.VoucherDate) AS VARCHAR)
              END
        JOIN MsItemMaster  im ON im.ItemID = vi.ItemID
        JOIN MsBrandMaster b  ON b.BrandID = im.BrandID
        WHERE h.TransTypeID IN ({type_ph}) AND h.Cancelled = 'N'
          AND h.VoucherDate >= DATEADD(MONTH, -{months_back}, GETDATE())
        GROUP BY b.BrandName, FORMAT(h.VoucherDate, 'yyyy-MM')
    """)
    if df.empty:
        return df
    df["Cases"] = pd.to_numeric(df["Cases"], errors="coerce").fillna(0.0)
    return df


def _family_match(brandname: str, patterns: list[str]) -> bool:
    bn = (brandname or "").upper()
    return any(p.upper() in bn for p in patterns)


def _month_iter(start: date, n: int) -> list[str]:
    """Return n consecutive 'yyyy-MM' strings starting at `start`."""
    out, y, m = [], start.year, start.month
    for _ in range(n):
        out.append(f"{y:04d}-{m:02d}")
        m += 1
        if m == 13:
            m = 1; y += 1
    return out


def _previous_month(yyyymm: str) -> str:
    y, m = map(int, yyyymm.split("-"))
    m -= 1
    if m == 0:
        m = 12; y -= 1
    return f"{y:04d}-{m:02d}"


def _section_target_phasing(stock_df: pd.DataFrame, as_of: date) -> None:
    st.subheader("🎯 Annual target phasing — Brindco transition (Jun-2026)")
    st.caption(
        "Brindco takes over these 6 Diageo brands from June. The annual "
        "case target gets split across 12 months using each brand's own "
        "historical seasonality — but **months with <1 case sold are treated "
        "as stock-out** (not zero demand) and replaced with the brand's "
        "active-month average. Month-1 indent is reduced by current stock."
    )

    # Family targets — editable per row
    init = pd.DataFrame([
        {"Code": p["code"], "Brand": p["name"],
         "Annual target (cs)": int(p["target"])}
        for p in BRINDCO_PRESET
    ])
    edited = st.data_editor(
        init, key="brindco_targets", use_container_width=True, hide_index=True,
        column_config={"Code": st.column_config.Column(disabled=True),
                       "Brand": st.column_config.Column(disabled=True),
                       "Annual target (cs)":
                           st.column_config.NumberColumn(min_value=0, step=10)},
        num_rows="fixed",
    )
    targets = {row["Code"]: float(row["Annual target (cs)"])
               for _, row in edited.iterrows()}

    monthly = _load_family_monthly(13)
    if monthly.empty:
        st.warning("No sales data available for the look-back window.")
        return

    # Family aggregation
    fam_codes  = [p["code"] for p in BRINDCO_PRESET]
    fam_meta   = {p["code"]: p for p in BRINDCO_PRESET}
    monthly["Code"] = monthly["BrandName"].map(
        lambda bn: next((c for c in fam_codes
                         if _family_match(bn, fam_meta[c]["patterns"])), None))
    fam = monthly.dropna(subset=["Code"])
    fam = fam.groupby(["Code", "Mon"], as_index=False)["Cases"].sum()

    # Stock per family from current stock_df
    stock_per_fam = {}
    if not stock_df.empty and "BrandName" in stock_df.columns:
        for c in fam_codes:
            mask = stock_df["BrandName"].apply(
                lambda bn: _family_match(bn, fam_meta[c]["patterns"]))
            stock_per_fam[c] = float(stock_df.loc[mask, "ClosingCases"].sum())
    else:
        stock_per_fam = {c: 0.0 for c in fam_codes}

    # L12M window = the 12 calendar months ending in the previous full month.
    # (today = May-2026 → L12M = May-2025 … Apr-2026.)
    cur_mon  = f"{as_of.year:04d}-{as_of.month:02d}"
    prev_mon = _previous_month(cur_mon)
    l12_months = []
    pm = prev_mon
    for _ in range(12):
        l12_months.append(pm); pm = _previous_month(pm)
    l12_months = list(reversed(l12_months))
    l3_months  = l12_months[-3:]

    summary_rows = []
    for c in fam_codes:
        f = fam[fam["Code"] == c].set_index("Mon")["Cases"]
        l12 = [float(f.get(m, 0.0)) for m in l12_months]
        l3  = [float(f.get(m, 0.0)) for m in l3_months]
        l12_active = sum(1 for v in l12 if v >= _STOCKOUT_THRESHOLD_CS)
        l3_active  = sum(1 for v in l3  if v >= _STOCKOUT_THRESHOLD_CS)
        l12_total  = sum(l12); l3_total = sum(l3)
        l12_naive  = l12_total / 12.0
        l12_adj    = l12_total / l12_active if l12_active else 0.0
        l3_naive   = l3_total  / 3.0
        l3_adj     = l3_total  / l3_active  if l3_active  else 0.0
        target     = targets.get(c, 0.0)
        target_mo  = target / 12.0
        # Headroom: ratio of target to the adjusted (real-demand) run rate
        headroom = (target_mo / l12_adj) if l12_adj > 0 else float("inf")
        summary_rows.append({
            "Code": c, "Brand": fam_meta[c]["name"],
            "Annual target":     target,
            "Target/mo":         target_mo,
            "L12M total":        l12_total,
            "L12M avg (naive)":  l12_naive,
            "L12M avg (adj)":    l12_adj,
            "L3M avg (naive)":   l3_naive,
            "L3M avg (adj)":     l3_adj,
            "Active mo (L12M)":  f"{l12_active}/12",
            "Stock now":         stock_per_fam.get(c, 0.0),
            "Target × adj":      headroom,
        })
    summary = pd.DataFrame(summary_rows)

    def _flag_headroom(v):
        try:
            if v == float("inf"):
                return "color:#dc2626;font-weight:700"   # no demand history
            if v > 1.5:
                return "color:#dc2626;font-weight:700"   # very ambitious
            if v > 1.0:
                return "color:#b45309;font-weight:600"   # ambitious
            return "color:#16a34a;font-weight:600"
        except Exception:
            return ""

    st.markdown("##### Summary — target vs your real history")
    money_cols = ["Annual target", "Target/mo", "L12M total",
                  "L12M avg (naive)", "L12M avg (adj)",
                  "L3M avg (naive)", "L3M avg (adj)", "Stock now"]
    st.dataframe(
        summary.style
               .format({c: "{:,.1f}" for c in money_cols}
                       | {"Target × adj": lambda v:
                          ("∞" if v == float("inf") else f"{v:,.2f}×")})
               .map(_flag_headroom, subset=["Target × adj"]),
        use_container_width=True, hide_index=True,
        column_config={
            "Target × adj": st.column_config.Column(
                "Target ÷ adj-avg",
                help=("Target monthly ÷ stock-out-adjusted L12M run-rate. "
                      ">1.5× = very ambitious; ∞ = no demand history.")),
        })

    # ── Phasing matrix ──
    # 12 forward months from next month.
    start_y = as_of.year + (1 if as_of.month == 12 else 0)
    start_m = 1 if as_of.month == 12 else as_of.month + 1
    forward = _month_iter(date(start_y, start_m, 1), 12)

    phasing_rows = []
    for c in fam_codes:
        f = fam[fam["Code"] == c].set_index("Mon")["Cases"]
        # Build weights for each forward month from same-month-last-year
        adj_avg = next((r["L12M avg (adj)"] for r in summary_rows
                        if r["Code"] == c), 0.0)
        active_l12 = sum(1 for m in l12_months
                         if float(f.get(m, 0.0)) >= _STOCKOUT_THRESHOLD_CS)
        flat_fallback = active_l12 < 4    # too little history to seasonally weight

        weights = []
        for fm in forward:
            # Same calendar month last year: "2026-06" → "2025-06"
            same_last_year = f"{int(fm[:4])-1:04d}-{fm[5:]}"
            hist = float(f.get(same_last_year, 0.0))
            if flat_fallback:
                w = 1.0
            elif hist >= _STOCKOUT_THRESHOLD_CS:
                w = hist
            else:
                w = adj_avg          # placeholder for a historical stock-out
            weights.append(w)
        wsum = sum(weights) or 1.0
        target = targets.get(c, 0.0)
        plan = [round(target * w / wsum, 1) for w in weights]
        # Reduce month-1 by current stock (floor 0)
        plan[0] = round(max(0.0, plan[0] - stock_per_fam.get(c, 0.0)), 1)

        row = {"Code": c, "Brand": fam_meta[c]["name"]}
        for m, v in zip(forward, plan):
            row[m] = v
        row["TOTAL"] = round(sum(plan), 1)
        if flat_fallback:
            row["Brand"] = row["Brand"] + " ⚠️ flat split (sparse history)"
        phasing_rows.append(row)

    phasing = pd.DataFrame(phasing_rows)
    st.markdown("##### Monthly indent plan — 12 months from next month")
    st.caption(
        "Each cell = recommended cases to indent. Month-1 already deducts your "
        "current stock. Brands flagged ⚠️ have <4 active months in L12M, so a "
        "flat 1/12 split is used instead of seasonality."
    )
    fmt = {c: "{:,.1f}" for c in forward + ["TOTAL"]}
    st.dataframe(
        phasing.style.format(fmt),
        use_container_width=True, hide_index=True)
    st.download_button(
        "⬇️ Download indent plan",
        phasing.to_csv(index=False).encode("utf-8-sig"),
        file_name=f"brindco_indent_{as_of.isoformat()}.csv",
        mime="text/csv", key="brindco_dl")


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN RENDER
# ═══════════════════════════════════════════════════════════════════════════════

def render() -> None:
    st.title("Inventory")

    sub = st.radio(
        "View",
        ["📊 Stock Analytics", "📦 Primary Plan / Indent", "📋 Brand Catalog"],
        horizontal=True, key="inv_sub_nav", label_visibility="collapsed")
    st.divider()

    today    = date.today()
    fy_start = date(today.year if today.month >= 4 else today.year - 1, 4, 1)

    c1, c2, c3 = st.columns([1, 2, 1])
    with c1:
        as_of_date = st.date_input(
            "Stock as of",
            value=today,
            key="inv_as_of",
        )
    with c2:
        principal_filter = st.multiselect(
            "Principal",
            options=list(_PRINCIPAL_NAMES.values()),
            default=[],
            key="inv_principal_filter",
        )
    with c3:
        view_filter = st.selectbox(
            "Show",
            ["All items", "Low stock (<10 cases)", "Out of stock", "Slow movers"],
            key="inv_view_filter",
        )

    keg_aware = keg_mode_toggle("inv_keg_mode", default_keg_aware=False)

    # Show what the stock is anchored to (ERP S&S baseline + live roll-forward)
    _bdate, _bitems = _load_stock_baseline()
    if _bdate:
        st.caption(
            f"📦 Stock = **FY-opening ({_bdate})** from the ERP Stock & Sale "
            f"report + live bill movements since. Computed automatically — "
            f"you only upload the S&S **once a year on 1 April** to set the new "
            f"FY opening. ({len(_bitems)} opening items.)"
        )
    else:
        st.caption("⚠️ No S&S baseline found — showing MsItemBatchOpening "
                   "(may not match the ERP Stock & Sale report).")

    # Reconciliation period (defaults to FY start through as_of).
    # Wrap the heavy loaders in an error boundary so a transient DB drop
    # (rare after the run_query retry tuning) shows a clean message + retry
    # hint instead of a raw pymssql traceback.
    try:
        with st.spinner("Loading stock + movement data…"):
            stock_df    = _build_stock_df(as_of_date)
            origin_df   = _load_item_origin()
            vel_30      = _load_sales_velocity(as_of_date, days_back=30, keg_aware=keg_aware)
            vel_90      = _load_sales_velocity(as_of_date, days_back=90, keg_aware=keg_aware)
    except Exception as exc:
        st.error(
            "⚠️ **The database connection dropped while loading stock data.**\n\n"
            "This is usually transient. Click **🔄 Refresh** at the top-right "
            "of the page (or hit Ctrl+Shift+R) to retry. If it keeps happening, "
            "the SQL Server may be under stress — try again in a minute.\n\n"
            f"_Technical details: `{type(exc).__name__}: {str(exc)[:200]}`_"
        )
        return    # exit render() cleanly — st.stop() would be belt-and-braces
                  # but in tests it's a no-op which leaks past unbound locals.

    if stock_df.empty:
        st.error("No stock data returned."); return

    # Apply keg mode to the stock case counts (value always uses the plain
    # physical-unit basis via ClosingCasesPlain, so it's unaffected).
    if not keg_aware:
        stock_df = stock_df.copy()
        stock_df["ClosingCases"] = stock_df["ClosingCasesPlain"]
        stock_df["CaseRem"]      = stock_df["ClosingCases"].astype(int)

    if principal_filter:
        stock_df = stock_df[stock_df["Principal"].isin(principal_filter)]

    if view_filter == "Low stock (<10 cases)":
        view_df = stock_df[stock_df["ClosingCases"] < 10]
    elif view_filter == "Out of stock":
        view_df = stock_df[stock_df["ClosingBottles"] <= 0]
    else:
        view_df = stock_df

    if as_of_date < today:
        st.info(f"Showing historical stock as of {as_of_date.strftime('%d %b %Y')} "
                f"(rolled back from live).")

    if sub.endswith("Stock Analytics"):
        safe_section("KPIs",            _section_kpis,            stock_df)
        st.divider()
        safe_section("Stock reconciliation", _section_reconciliation,
                     stock_df, keg_aware)
        st.divider()
        safe_section("By principal",    _section_by_principal,    stock_df)
        st.divider()
        safe_section("Top items",       _section_top_items,       view_df, origin_df)
        st.divider()
        safe_section("Slow movers",     _section_slow_movers,     stock_df, vel_30)
        st.divider()
        safe_section("Out of stock",    _section_out_of_stock,    stock_df, vel_30, vel_90)
        st.divider()
        safe_section("Days of cover",   _section_days_of_cover,   stock_df, vel_30)
        st.divider()
        safe_section("Indent planner",  _section_indent_planner,  stock_df, as_of_date, keg_aware)
    elif sub.endswith("Primary Plan / Indent"):
        safe_section("Primary Plan / Indent",
                     _section_primary_plan, stock_df, as_of_date)
    else:
        safe_section("Brand Catalog", _section_brand_catalog, stock_df)
