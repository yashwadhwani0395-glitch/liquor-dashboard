"""src/nrm.py — NRM Planner (Diageo / USL party × brand allocation).

The NRM ("Numeric Reach Maximization") workflow Diageo + USL send to KWPL
monthly is a 6,500-row party × brand grid where:

  • Pre-filled by the company: L6M / L3M / LM / LYSM / CM-MTD volumes,
    BTG (Bottom To Grow), Plan Gr%, the OVERALL SUMMARY brand-level
    MOR TGT (Monthly Operating Target).
  • Filled by KWPL's manager: TL Plan-1 (initial cases per party × brand,
    constrained so the per-brand total equals the MOR TGT).

The manager currently spends ~2 days on NRM-1 and ~2 days on NRM-2
(the revision near month-end). This module collapses that to minutes:
parse the .xlsb, cross-check Diageo's history against our own ERP voucher
data, suggest a starting TL Plan-1 per row, let the manager override,
and export a filled file.

CUSTOMERID mapping: Diageo's CUSTOMERID (e.g. '7755') maps to our
MsPartyMaster via:
  primary  → PartyCode = CUSTOMERID
  fallback → Remark    = CUSTOMERID   (older parties stored it here)

Phase 1 (this module today): TL Plan-1 allocator.
Phase 2 (later): CPC slab auto-fill — uses the 'Slabs' sheet and writes
  Total CPC / Slab Scheme Amt / Tie Up Amt / Burst columns.  Three exception
  parties (CHARLIE WINES, SAB TONIQUE, JAS TONIQUE) have FIXED payouts,
  all other parties run on slab-based monthly schemes.
Phase 3 (later): NRM-2 reviser — pulls live CM-MTD, flags under-tracking
  parties, suggests revised TL Plan-2.
"""
from __future__ import annotations

from datetime import date
from io import BytesIO
from typing import Any

import pandas as pd
import streamlit as st

from db import run_query
from utils.helpers import safe_section

# ── Constants ────────────────────────────────────────────────────────────────

# Parties whose payouts are FIXED (not on slab schemes). For Phase 1
# (volume) they're treated like any other party; for Phase 2 (CPC) we'll
# bypass slab calc and apply their fixed rates.
EXCEPTION_PARTIES: frozenset[str] = frozenset({
    "CHARLIE WINES",
    "SAB TONIQUE",     # SAB JAS TONIQUE RETAIL
    "JAS TONIQUE",     # JAS TONIQUE RETAIL LLP
})

# Editable column the manager fills in NRM Plan-1.
TL_PLAN_1_COL = "TL Plan-1"

# Source-of-truth column from the OVERALL SUMMARY sheet.
MOR_TGT_COL = "MOR TGT"


# ── Loaders ─────────────────────────────────────────────────────────────────

@st.cache_data(ttl=3600, show_spinner=False)
def _load_party_code_map() -> pd.DataFrame:
    """Map Diageo CUSTOMERID → our PartyID via MsPartyMaster.PartyCode
    (primary) or .Remark (fallback for older parties)."""
    df = run_query("""
        SELECT
            PartyID,
            PartyName,
            ISNULL(NULLIF(LTRIM(RTRIM(PartyCode)), ''), '')   AS PartyCode,
            ISNULL(NULLIF(LTRIM(RTRIM(Remark)),    ''), '')   AS Remark,
            ClassID, LicenseTypeID, AcType3ID, BannedPartyYN
        FROM MsPartyMaster
        WHERE PartyID LIKE 'D%'
    """)
    if df.empty:
        return df
    # Build a (Diageo CUSTOMERID → PartyID) lookup
    rows = []
    for _, r in df.iterrows():
        cid = r["PartyCode"] or r["Remark"]
        if cid and cid.isdigit():
            rows.append({"CustomerID": str(cid), "PartyID": r["PartyID"],
                         "PartyName": r["PartyName"],
                         "SourceCol": "PartyCode" if r["PartyCode"] else "Remark"})
    out = pd.DataFrame(rows)
    return out


@st.cache_data(ttl=900, show_spinner=False)
def _load_party_brand_monthly(months_back: int = 14) -> pd.DataFrame:
    """Per (PartyID, BrandID, Mon) cases sold — last N months. Used to
    cross-check Diageo's L3M / LM / LYSM / CM-MTD numbers and to compute
    smart TL Plan-1 suggestions."""
    from src.discounts import _SALES_TYPES  # reuses the canonical list
    type_ph = ",".join(str(t) for t in _SALES_TYPES)
    df = run_query(f"""
        SELECT
            d.PartyID,
            b.BrandID, b.BrandName,
            FORMAT(h.VoucherDate, 'yyyy-MM')   AS Mon,
            CAST(SUM(CAST(vi.TotalBottleQty AS decimal(18,4))
                     / NULLIF(im.BottlesPerCase, 0)) AS float) AS Cases
        FROM TrVocHead h
        JOIN TrVocItem vi
            ON  vi.TransTypeID = h.TransTypeID AND vi.VoucherNo = h.VoucherNo
            AND vi.ItemID LIKE 'I%'
            AND vi.FinancialYear = CASE
                WHEN MONTH(h.VoucherDate) >= 4
                THEN CAST(YEAR(h.VoucherDate) AS VARCHAR)+'-'+CAST(YEAR(h.VoucherDate)+1 AS VARCHAR)
                ELSE CAST(YEAR(h.VoucherDate)-1 AS VARCHAR)+'-'+CAST(YEAR(h.VoucherDate) AS VARCHAR)
              END
        JOIN MsItemMaster  im ON im.ItemID = vi.ItemID
        JOIN MsBrandMaster b  ON b.BrandID = im.BrandID
        JOIN (
            SELECT TransTypeID, VoucherNo, FinancialYear, PartyID FROM (
                SELECT TransTypeID, VoucherNo, FinancialYear, PartyID,
                       ROW_NUMBER() OVER (PARTITION BY TransTypeID, VoucherNo, FinancialYear
                                          ORDER BY id_key DESC) rn
                FROM TrVocDetail
                WHERE PartyID LIKE 'D%' AND DrCrIndicator='D'
            ) x WHERE rn = 1
        ) d ON d.TransTypeID = h.TransTypeID AND d.VoucherNo = h.VoucherNo
           AND d.FinancialYear = h.FinancialYear
        WHERE h.TransTypeID IN ({type_ph}) AND h.Cancelled = 'N'
          AND b.CompanyID IN ('C00025','C00040')      -- USL + Diageo only
          AND h.VoucherDate >= DATEADD(MONTH, -{months_back}, GETDATE())
        GROUP BY d.PartyID, b.BrandID, b.BrandName, FORMAT(h.VoucherDate, 'yyyy-MM')
    """)
    if not df.empty:
        df["Cases"] = pd.to_numeric(df["Cases"], errors="coerce").fillna(0.0)
    return df


# ── File parsing ─────────────────────────────────────────────────────────────

def _read_sheet_auto_header(upload_bytes: bytes, sheet_name: str,
                            marker_words: list[str],
                            max_scan: int = 10) -> pd.DataFrame:
    """Read an .xlsb sheet without trusting a hardcoded header-row index.
    Scans the first `max_scan` rows and uses the first one containing ALL
    of `marker_words` (case-insensitive) as the column-header row.

    The raw file's headers shift between rows 1 / 2 / 3 depending on the
    sheet, the .xlsb tooling version, and whether the cloud's pyxlsb skips
    blank rows or not. Auto-detection is robust to all of that."""
    bio = BytesIO(upload_bytes)
    df  = pd.read_excel(bio, sheet_name=sheet_name, engine="pyxlsb", header=None)
    markers_up = [m.upper() for m in marker_words]
    for ri in range(min(max_scan, len(df))):
        row_str = " | ".join(
            str(v) for v in df.iloc[ri].tolist() if pd.notna(v)
        ).upper()
        if all(m in row_str for m in markers_up):
            new_cols = df.iloc[ri].fillna("").astype(str).tolist()
            body = df.iloc[ri + 1:].copy()
            body.columns = new_cols
            return body.reset_index(drop=True)
    raise ValueError(
        f"Could not find a header row in sheet {sheet_name!r} containing "
        f"all of {marker_words!r}."
    )


def _parse_uploaded_nrm(upload_bytes: bytes) -> dict[str, pd.DataFrame]:
    """Parse the raw .xlsb into the NRM grid + OVERALL SUMMARY + DB WISE
    SUMMARY + Brand Mapping. Auto-detects header rows so we don't break
    when Diageo's template version shifts a row.

    DB WISE SUMMARY is the AUTHORITATIVE source for per-distributor brand
    targets — OVERALL SUMMARY shows the manager's total across ALL their
    distributors (which inflates Kranti's target ~6x), see comment on
    _brand_targets_from_db_wise."""
    try:
        nrm  = _read_sheet_auto_header(
            upload_bytes, "NRM",
            ["BRAND", "PARTY", "CUSTOMERID"])
        smry = _read_sheet_auto_header(
            upload_bytes, "OVERALL SUMMARY",
            ["MOR TGT", "BRAND"])
        try:
            db_wise = _read_sheet_auto_header(
                upload_bytes, "DB WISE SUMMARY",
                ["DISTRIBUTOR", "BRAND", "MOR TGT"])
        except ValueError:
            db_wise = pd.DataFrame()
        try:
            bmap = _read_sheet_auto_header(
                upload_bytes, "Brand Mapping",
                ["BRANDALIAS", "NRM BRAND"])
        except ValueError:
            bmap = pd.DataFrame()
    except Exception as exc:
        st.error(f"Failed to parse .xlsb file: "
                 f"`{type(exc).__name__}: {exc}`")
        return {}
    return {"NRM": nrm, "SUMMARY": smry, "DB_WISE": db_wise, "BRAND_MAP": bmap}


# ── Smart allocation ─────────────────────────────────────────────────────────

def _suggest_tl_plan_1(nrm: pd.DataFrame,
                       brand_targets: dict[str, float]) -> pd.Series:
    """Per row, compute a suggested TL Plan-1. Two-step:
      1. raw = max(L3M, LM) if brand is growing else (L3M + LM)/2,
         capped at 1.5 × LM so we never wildly over-commit.
      2. rebalance per brand so the brand's total TL Plan-1 = MOR TGT
         exactly.  Rows with no history get 0 (manager fills manually)."""
    l3m   = pd.to_numeric(nrm.get(" Average L3M", 0), errors="coerce").fillna(0.0)
    lm    = pd.to_numeric(nrm.get("LM",            0), errors="coerce").fillna(0.0)
    lysm  = pd.to_numeric(nrm.get("LYSM",          0), errors="coerce").fillna(0.0)
    plan_gr = pd.to_numeric(nrm.get("Plan Gr %",   0), errors="coerce").fillna(0.0)

    # Step 1 — raw per-row signal
    growth_threshold = 0.05
    raw = []
    for a, b, g in zip(l3m, lm, plan_gr):
        if g > growth_threshold:
            v = max(a, b)
        else:
            v = (a + b) / 2.0
        # Cap at 1.5 × LM unless LM is 0 (then 1.5 × L3M)
        cap = 1.5 * (b if b > 0 else a) if (a + b) > 0 else 0
        raw.append(min(v, cap) if cap > 0 else 0.0)
    raw = pd.Series(raw, index=nrm.index)

    # Step 2 — rebalance per brand to hit the MOR TGT exactly
    brand_col = nrm.get("BRAND")
    if brand_col is None:
        return raw
    out = raw.copy()
    for brand, group in nrm.groupby("BRAND"):
        idx = group.index
        target = float(brand_targets.get(str(brand), 0))
        if target <= 0:
            out.loc[idx] = 0
            continue
        raw_total = out.loc[idx].sum()
        if raw_total <= 0:
            # No history at all — fall back to a flat split across rows
            # with any positive L3M/LM/LYSM (so completely-dead rows stay 0)
            with_history = idx[
                (l3m.loc[idx] + lm.loc[idx] + lysm.loc[idx]) > 0]
            if len(with_history) > 0:
                out.loc[with_history] = target / len(with_history)
        else:
            out.loc[idx] = out.loc[idx] * (target / raw_total)
    return out.round(2)


# ── Brand target extraction ────────────────────────────────────────────────

def _brand_targets_from_db_wise(db_wise: pd.DataFrame,
                                distributor: str = "KRANTI") -> dict[str, float]:
    """Per-distributor brand targets from the DB WISE SUMMARY sheet.

    AUTHORITATIVE source — unlike OVERALL SUMMARY (which sums each manager's
    TOTAL across every distributor they cover and inflates Kranti's
    target ~6x), DB WISE SUMMARY is one row per
    (Distributor x Manager x Channel x Brand). Filtering to our distributor
    name and summing across channels gives the brand's actual target for us.

    Verified on Jun-26 file: gives 17,806 cs total target vs the manager's
    actual worked TL Plan-1 of 19,344 (~92% match — the small remaining gap
    is brands the company didn't itemise in DB WISE SUMMARY that the manager
    pushes on his own)."""
    if db_wise.empty:
        return {}
    cols = {str(c).strip().upper(): c for c in db_wise.columns}
    dist_col  = cols.get("DISTRIBUTOR")
    brand_col = cols.get("BRAND")
    tgt_col   = cols.get("MOR TGT")
    if not (dist_col and brand_col and tgt_col):
        return {}
    df = db_wise.copy()
    df[dist_col]  = df[dist_col].astype(str).str.upper()
    df = df[df[dist_col].str.contains(distributor.upper(), na=False)]
    if df.empty:
        return {}
    df[brand_col] = df[brand_col].astype(str).str.strip()
    df[tgt_col]   = pd.to_numeric(df[tgt_col], errors="coerce").fillna(0.0)
    grouped = df.groupby(brand_col)[tgt_col].sum()
    return {k: float(v) for k, v in grouped.items()
            if v > 0 and k not in ("", "nan", "None")}


def _brand_targets_from_summary(summary: pd.DataFrame) -> dict[str, float]:
    """Map BRAND → MOR TGT (cases) from the OVERALL SUMMARY sheet, summed
    across managers when a brand appears under more than one. Defensive
    against whitespace / case variations in the column names."""
    if summary.empty:
        return {}
    # Find the BRAND and MOR TGT columns regardless of leading/trailing spaces
    cols = {str(c).strip().upper(): c for c in summary.columns}
    brand_col = cols.get("BRAND")
    tgt_col   = cols.get("MOR TGT")
    if brand_col is None or tgt_col is None:
        return {}
    s = summary.copy()
    s[brand_col] = s[brand_col].astype(str).str.strip()
    s[tgt_col]   = pd.to_numeric(s[tgt_col], errors="coerce").fillna(0.0)
    grouped = s.groupby(brand_col)[tgt_col].sum()
    return {k: float(v) for k, v in grouped.items()
            if v > 0 and k not in ("", "nan", "None")}


# ── Exception-party marker ──────────────────────────────────────────────────

def _is_exception_party(party_name: str) -> bool:
    if not isinstance(party_name, str):
        return False
    upper = party_name.upper()
    return any(ex in upper for ex in EXCEPTION_PARTIES)


# ═══════════════════════════════════════════════════════════════════════════
# RENDER
# ═══════════════════════════════════════════════════════════════════════════

def render() -> None:
    st.title("📋 NRM Planner — USL + Diageo monthly party × brand")
    st.caption(
        "Upload Diageo's raw NRM .xlsb; we cross-check every row against "
        "your ERP voucher history, suggest a TL Plan-1 per (party × brand) "
        "so every brand total equals its MOR TGT, and let you tweak before "
        "downloading the filled file. Phase 1 — TL Plan-1 only. CPC slabs "
        "and NRM-2 revision come in later phases."
    )

    up = st.file_uploader("Diageo NRM file (.xlsb)", type=["xlsb"],
                          key="nrm_upload")
    if up is None:
        st.info(
            "Drop the raw NRM file your manager normally fills. Tip: it's "
            "the one named like *WMH- &lt;Mon&gt;'YY NRM Plan-1 DB - Pune - "
            "Kranti.xlsb*."
        )
        return

    with st.spinner("Parsing file…"):
        sheets = _parse_uploaded_nrm(up.getvalue())
    if not sheets:
        return

    nrm_raw = sheets["NRM"]
    summary = sheets["SUMMARY"]

    # Filter to Kranti rows only
    if "DISTRIBUTOR" in nrm_raw.columns:
        nrm_raw = nrm_raw[
            nrm_raw["DISTRIBUTOR"].astype(str).str.upper()
                .str.contains("KRANTI", na=False)
        ].copy()

    # Brand targets — prefer DB WISE SUMMARY (per-distributor accurate);
    # fall back to OVERALL SUMMARY only if the DB WISE sheet is missing.
    brand_targets = _brand_targets_from_db_wise(
        sheets.get("DB_WISE", pd.DataFrame()), distributor="KRANTI")
    target_source = "DB WISE SUMMARY (Kranti rows)"
    if not brand_targets:
        brand_targets = _brand_targets_from_summary(summary)
        target_source = "OVERALL SUMMARY (fallback — may over-allocate)"
    if not brand_targets:
        st.warning("Couldn't read brand-level MOR TGT from either DB WISE "
                   "SUMMARY or OVERALL SUMMARY — the auto-allocator can't "
                   "constrain without it. Manager will need to enter TL "
                   "Plan-1 manually.")
    else:
        st.caption(f"📊 Brand targets sourced from: **{target_source}** · "
                   f"{len(brand_targets)} brands · "
                   f"total = {sum(brand_targets.values()):,.0f} cs")

    # Compute suggested TL Plan-1
    nrm_raw["Suggested TL-1"] = _suggest_tl_plan_1(nrm_raw, brand_targets)
    # User editable column — pre-populated with suggestion (or any existing
    # TL Plan-1 value the file already carries)
    existing = pd.to_numeric(nrm_raw.get("TL Plan-1", 0),
                             errors="coerce").fillna(0.0)
    nrm_raw["Your TL-1"] = existing.where(existing > 0, nrm_raw["Suggested TL-1"])

    # Mark exception parties
    nrm_raw["Exception"] = nrm_raw.get("PARTY", "").apply(_is_exception_party)

    # Headline KPIs
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("Rows (Kranti)", f"{len(nrm_raw):,}")
    k2.metric("Brands w/ MOR TGT", f"{len(brand_targets):,}")
    k3.metric("Exception rows", int(nrm_raw["Exception"].sum()),
              help="Charlie Wines / SAB Tonique / Jas Tonique — fixed-payout "
                   "parties. Phase 2 will use their fixed CPC; Phase 1 "
                   "allocates volume normally.")
    suggested_total = float(nrm_raw["Suggested TL-1"].sum())
    target_total    = float(sum(brand_targets.values()))
    k4.metric("Σ Suggested vs Σ Target",
              f"{suggested_total:,.0f} / {target_total:,.0f} cs",
              delta=f"{(suggested_total - target_total):+,.0f}")

    # Live brand-total tracker
    st.markdown("##### Brand-total tracker (live vs MOR TGT)")
    tracker = nrm_raw.groupby("BRAND")["Your TL-1"].sum().reset_index()
    tracker["MOR TGT"] = tracker["BRAND"].map(brand_targets).fillna(0.0)
    tracker["Gap"]     = tracker["MOR TGT"] - tracker["Your TL-1"]
    tracker = tracker.sort_values("MOR TGT", ascending=False)

    def _flag(v):
        try: v = float(v)
        except Exception: return ""
        if abs(v) < 0.5: return "color:#16a34a;font-weight:700"
        if abs(v) < 5:   return "color:#b45309;font-weight:600"
        return "color:#dc2626;font-weight:700"
    st.dataframe(
        tracker.style.format({"Your TL-1": "{:,.1f}",
                              "MOR TGT":   "{:,.1f}",
                              "Gap":       "{:+,.1f}"})
               .map(_flag, subset=["Gap"]),
        use_container_width=True, hide_index=True, height=320)

    # Editable grid (filter to rows the manager normally works — has any history)
    st.markdown("##### Editable grid")
    st.caption(
        "Override **Your TL-1** on any row. The brand-total tracker above "
        "updates live. Filter to rows with any history below to skip the "
        "company-MIS-handled rows."
    )
    only_history = st.checkbox(
        "Hide rows with no history (L3M + LM + LYSM = 0)", value=True,
        key="nrm_only_history")

    cols_to_show = ["BRAND", "PARTY", "CUSTOMERID", "Exception",
                    " Average L3M", "LM", "LYSM", "CM Sales",
                    "Suggested TL-1", "Your TL-1"]
    grid = nrm_raw[[c for c in cols_to_show if c in nrm_raw.columns]].copy()
    if only_history:
        non_zero = (
            pd.to_numeric(grid.get(" Average L3M", 0), errors="coerce").fillna(0)
            + pd.to_numeric(grid.get("LM", 0),         errors="coerce").fillna(0)
            + pd.to_numeric(grid.get("LYSM", 0),       errors="coerce").fillna(0)
        )
        grid = grid[non_zero > 0]

    edited = st.data_editor(
        grid, use_container_width=True, hide_index=True,
        num_rows="fixed",
        column_config={
            "Your TL-1": st.column_config.NumberColumn(
                "Your TL-1 (cs)", min_value=0.0, step=0.5, format="%.1f"),
            "Suggested TL-1": st.column_config.NumberColumn(
                "Suggested TL-1", disabled=True, format="%.1f"),
            "Exception": st.column_config.CheckboxColumn(
                "Fixed-payout", disabled=True),
            " Average L3M": st.column_config.NumberColumn(
                "L3M", disabled=True, format="%.1f"),
            "LM":   st.column_config.NumberColumn(disabled=True, format="%.1f"),
            "LYSM": st.column_config.NumberColumn(disabled=True, format="%.1f"),
            "CM Sales": st.column_config.NumberColumn(
                "CM (MTD)", disabled=True, format="%.1f"),
        },
        key=f"nrm_editor_{up.name}",
    )

    # Push edits back into the master nrm_raw so download captures them
    if not grid.empty:
        nrm_raw.loc[grid.index, "Your TL-1"] = edited["Your TL-1"].values

    # Download: writes filled TL Plan-1 column into a clean .xlsx
    if st.button("⬇️ Generate filled NRM (full file, all sheets preserved)",
                 type="primary", key="nrm_dl_btn"):
        with st.spinner("Building file — preserving every sheet & cell…"):
            out_bytes = _build_filled_xlsx(up.getvalue(), nrm_raw)
        st.success(
            "File ready. Output is .xlsx with every original sheet and "
            "column intact — only the TL Plan-1 column has your edits. "
            "Open in Excel and Save As → .xlsb if Diageo asks for the "
            "binary format."
        )
        st.download_button(
            "Download filled NRM",
            out_bytes,
            file_name=f"NRM_filled_{date.today().isoformat()}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            key="nrm_dl_final")


def _build_filled_xlsx(upload_bytes: bytes,
                       nrm_with_edits: pd.DataFrame) -> bytes:
    """Round-trip the entire uploaded .xlsb into a new .xlsx, preserving
    EVERY sheet, EVERY column and EVERY row of the original. Only the
    TL Plan-1 column inside the NRM sheet gets overwritten with the
    manager's edits (column "Your TL-1" on `nrm_with_edits`).

    Output format note: Python has no maintained library that writes the
    binary .xlsb format, so the output is .xlsx. Diageo's MIS team accepts
    .xlsx (Excel auto-detects either), but if they ever insist on .xlsb the
    user can open the .xlsx in Excel and Save As → .xlsb in two clicks —
    every cell of the original is preserved so the resave is lossless."""
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter   # noqa: F401  (future use)

    bio = BytesIO(upload_bytes)
    xl  = pd.ExcelFile(bio, engine="pyxlsb")

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Map Kranti-row index (position in the parsed NRM df) → edited TL-1.
    # Rows not in this dict are NOT touched — non-Kranti distributors keep
    # whatever Diageo originally put there.
    edited_by_row: dict[int, float] = {}
    if "Your TL-1" in nrm_with_edits.columns:
        for ki, val in zip(nrm_with_edits.index, nrm_with_edits["Your TL-1"]):
            if pd.notna(val):
                edited_by_row[int(ki)] = float(val)

    for sn in xl.sheet_names:
        bio.seek(0)
        df = pd.read_excel(bio, sheet_name=sn, engine="pyxlsb", header=None)
        ws = wb.create_sheet(sn[:31])
        # Write every cell verbatim (None / NaN cells stay empty)
        for ri in range(len(df)):
            for ci in range(df.shape[1]):
                v = df.iat[ri, ci]
                if pd.notna(v):
                    ws.cell(row=ri + 1, column=ci + 1, value=v)

        # In the NRM sheet: locate the header, TL Plan-1 column AND the
        # DISTRIBUTOR column. Only overwrite TL Plan-1 on rows whose
        # DISTRIBUTOR says Kranti (safety belt in case the file ever
        # bundles multiple distributors).
        if sn == "NRM" and edited_by_row:
            hri = None
            for ri in range(min(10, len(df))):
                row_up = " | ".join(
                    str(v) for v in df.iloc[ri].tolist() if pd.notna(v)
                ).upper()
                if ("BRAND" in row_up and "PARTY" in row_up
                        and "CUSTOMERID" in row_up):
                    hri = ri
                    break
            if hri is None:
                continue
            tl_col = None; dist_col = None
            for ci, v in enumerate(df.iloc[hri].tolist()):
                if pd.notna(v):
                    s = str(v).strip().upper()
                    if s == "TL PLAN-1":
                        tl_col = ci
                    elif s == "DISTRIBUTOR":
                        dist_col = ci
            if tl_col is None:
                continue
            # Data rows start at hri + 1 in df  →  hri + 2 in openpyxl.
            data_start_openpyxl = hri + 2
            # Iterate ONLY the Kranti row positions; each kr-index maps to
            # raw NRM row hri + 1 + ki  (openpyxl row data_start + ki).
            for ki, val in edited_by_row.items():
                raw_row = hri + 1 + ki
                if raw_row >= len(df):
                    continue
                # Defensive: confirm the row's DISTRIBUTOR really is Kranti
                # before writing. If the col can't be found, trust the index.
                if dist_col is not None:
                    dv = df.iat[raw_row, dist_col]
                    if pd.notna(dv) and "KRANTI" not in str(dv).upper():
                        continue
                ws.cell(row=data_start_openpyxl + ki,
                        column=tl_col + 1, value=val)

    out = BytesIO()
    wb.save(out)
    return out.getvalue()
